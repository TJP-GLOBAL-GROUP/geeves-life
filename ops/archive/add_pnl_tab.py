import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

# Open the latest workbook
wb = openpyxl.load_workbook('/home/ubuntu/tax_prep/Tarik_Perkins_2025_Property_Tax_Workbook_v6.xlsx')

# Create new P&L sheet
if 'P&L by Property (2024)' in wb.sheetnames:
    del wb['P&L by Property (2024)']

ws = wb.create_sheet('P&L by Property (2024)')

# Styles
header_font = Font(bold=True, size=12)
title_font = Font(bold=True, size=14)
section_font = Font(bold=True, size=11)
currency_fmt = '#,##0.00;(#,##0.00);"-"'
border_bottom = Border(bottom=Side(style='thin'))
border_top_bottom = Border(top=Side(style='thin'), bottom=Side(style='double'))

green_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
red_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
blue_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
header_font_white = Font(bold=True, size=11, color='FFFFFF')

# Column widths
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 14

# Title
ws['A1'] = '2024 Short Term Rental P&L by Property'
ws['A1'].font = title_font
ws['A2'] = 'Allocation Method: 50/50 for NY shared expenses'
ws['A2'].font = Font(italic=True, size=10)

# Headers
row = 4
headers = ['', 'Penthouse', 'Sunset Studio', "Artiste's Boutique", 'TOTAL']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

# Income section
row = 6
ws.cell(row=row, column=1, value='INCOME').font = section_font
ws.cell(row=row, column=1).fill = green_fill
for col in range(2, 6):
    ws.cell(row=row, column=col).fill = green_fill

income_data = [
    ('Airbnb', 12457.65, 5472.30, 34442.73, 52372.68),
    ('VRBO', 9759.53, 5100.12, 667.00, 15526.65),
    ('Booking.com', 0, 0, 2105.38, 2105.38),
    ('PayPal (Booking.com guests)', 0, 0, 1637.91, 1637.91),
]

for item in income_data:
    row += 1
    ws.cell(row=row, column=1, value=item[0])
    for col in range(2, 6):
        val = item[col-1]
        cell = ws.cell(row=row, column=col, value=val if val != 0 else None)
        cell.number_format = currency_fmt
        cell.alignment = Alignment(horizontal='right')

# Total Income row
row += 1
ws.cell(row=row, column=1, value='Total Income').font = Font(bold=True)
totals_income = [22217.18, 10572.42, 38853.02, 71642.62]
for col, val in enumerate(totals_income, 2):
    cell = ws.cell(row=row, column=col, value=val)
    cell.number_format = currency_fmt
    cell.font = Font(bold=True)
    cell.border = border_bottom
    cell.alignment = Alignment(horizontal='right')

# Blank row
row += 2

# Expenses section
ws.cell(row=row, column=1, value='EXPENSES').font = section_font
ws.cell(row=row, column=1).fill = red_fill
for col in range(2, 6):
    ws.cell(row=row, column=col).fill = red_fill

expense_data = [
    ('Repairs & Maintenance', -14485.48, -14485.48, 0, -28970.96),
    ('Utilities (Electric)', -2589.50, -2589.50, -17597.88, -22776.89),
    ('Cleaning', -1527.50, -1527.50, -2518.33, -5573.33),
    ('Insurance', -1244.67, -1244.67, -117.83, -2607.17),
    ('Mortgage Interest', -1316.66, -1316.66, 0, -2633.33),
    ('Taxes & Licenses', -749.11, -749.11, 0, -1498.21),
    ('Administrative', -297.50, -297.50, 0, -595.00),
    ('Supplies & Furnishings', 201.68, 201.68, 0, 403.36),
    ('Legal & Professional', -195.06, -195.06, 0, -390.13),
    ('Advertising', 0, 0, 261.64, 261.64),
]

for item in expense_data:
    row += 1
    ws.cell(row=row, column=1, value=item[0])
    for col in range(2, 6):
        val = item[col-1]
        cell = ws.cell(row=row, column=col, value=val if val != 0 else None)
        cell.number_format = currency_fmt
        cell.alignment = Alignment(horizontal='right')

# Total Expenses row
row += 1
ws.cell(row=row, column=1, value='Total Expenses').font = Font(bold=True)
totals_expenses = [-22203.81, -22203.81, -19972.40, -64380.02]
for col, val in enumerate(totals_expenses, 2):
    cell = ws.cell(row=row, column=col, value=val)
    cell.number_format = currency_fmt
    cell.font = Font(bold=True)
    cell.border = border_bottom
    cell.alignment = Alignment(horizontal='right')

# Blank row
row += 2

# NET P&L
ws.cell(row=row, column=1, value='NET P&L').font = Font(bold=True, size=12)
ws.cell(row=row, column=1).fill = blue_fill
for col in range(2, 6):
    ws.cell(row=row, column=col).fill = blue_fill

net_pnl = [13.37, -11631.39, 18880.62, 7262.60]
for col, val in enumerate(net_pnl, 2):
    cell = ws.cell(row=row, column=col, value=val)
    cell.number_format = currency_fmt
    cell.font = Font(bold=True, size=12)
    cell.border = border_top_bottom
    cell.alignment = Alignment(horizontal='right')
    cell.fill = blue_fill

# Notes section
row += 3
ws.cell(row=row, column=1, value='NOTES:').font = Font(bold=True)
row += 1
ws.cell(row=row, column=1, value='• Income is on accrual basis (verified from Airbnb/VRBO/Booking.com exports)')
row += 1
ws.cell(row=row, column=1, value='• NY shared expenses allocated 50/50 between Penthouse and Sunset Studio')
row += 1
ws.cell(row=row, column=1, value='• Jamaica utilities include JPS electric ($14,302) + Boss Revolution telecom ($456)')
row += 1
ws.cell(row=row, column=1, value='• Repairs & Maintenance includes $25,547 in labor - review for capital improvement classification')
row += 1
ws.cell(row=row, column=1, value='• Mortgage Interest is to Jim Curatolo (private lender)')
row += 1
ws.cell(row=row, column=1, value='• Advertising shows net positive due to Booking.com refund ($570) offsetting commission ($308)')

# Save
output_path = '/home/ubuntu/tax_prep/Tarik_Perkins_2025_Property_Tax_Workbook_v7.xlsx'
wb.save(output_path)
print(f"Saved workbook with P&L tab: {output_path}")
print(f"Sheets: {wb.sheetnames}")
