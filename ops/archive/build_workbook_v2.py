"""
Build the 2025 US Property Tax Preparation Workbook v2.
Now uses FULL reservation confirmation data with complete financial breakdowns.
Sheets:
1. US Income Dashboard - Monthly pivot by property with full financials
2. US Booking Details - Every booking with guest paid, commission, payout
3. US Expenses - Categorized property-related expenses
4. US P&L Summary - Income minus expenses per property (Schedule E ready)
5. Jamaica Income - Jamaica property income (Artiste's Boutique)
6. Jamaica Booking Details - Full Jamaica booking breakdown
"""
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# Load parsed data
with open('parsed/airbnb_reservations_full.json') as f:
    all_reservations = json.load(f)

with open('parsed/property_expenses_parsed.json') as f:
    property_expenses = json.load(f)

# ============================================================
# NORMALIZE PROPERTY NAMES
# ============================================================
def normalize_property(name):
    if not name:
        return 'Unknown'
    name_lower = name.lower()
    if 'morabeza' in name_lower or 'seneca haven' in name_lower:
        return 'Morabeza (Studio)'
    elif 'sunset' in name_lower or 'seneca sunsets' in name_lower:
        return 'Sunset Studio'
    elif 'artis' in name_lower or 'bohemian' in name_lower or 'boutique' in name_lower:
        return "The Artiste's Boutique"
    return name

def get_month(date_str):
    if not date_str or len(date_str) < 7:
        return None
    try:
        return int(date_str[5:7])
    except:
        return None

def is_jamaica_property(prop_name):
    return prop_name == "The Artiste's Boutique"

# Separate 2025 and 2026 bookings
res_2025 = [r for r in all_reservations if r.get('check_in_date','').startswith('2025') and 'error' not in r]
res_2026 = [r for r in all_reservations if r.get('check_in_date','').startswith('2026') and 'error' not in r]

# Normalize property names
for r in res_2025:
    r['_norm_property'] = normalize_property(r.get('property_name', ''))
for r in res_2026:
    r['_norm_property'] = normalize_property(r.get('property_name', ''))

# Separate US and Jamaica for 2025
us_bookings_2025 = [r for r in res_2025 if not is_jamaica_property(r['_norm_property'])]
jamaica_bookings_2025 = [r for r in res_2025 if is_jamaica_property(r['_norm_property'])]

# Also include 2026 Jamaica bookings (future) for reference
jamaica_bookings_future = [r for r in res_2026 if is_jamaica_property(r['_norm_property'])]

print(f"2025 US bookings: {len(us_bookings_2025)}")
print(f"2025 Jamaica bookings: {len(jamaica_bookings_2025)}")
print(f"2026 future bookings: {len(res_2026)}")

# ============================================================
# BUILD WORKBOOK
# ============================================================
wb = Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
subheader_font = Font(bold=True, size=11)
subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
money_format = '#,##0.00'
pct_format = '0.0%'
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def auto_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 35)

# ============================================================
# SHEET 1: US Income Dashboard
# ============================================================
ws1 = wb.active
ws1.title = "US Income Dashboard"

ws1['A1'] = "2025 US Property Income Dashboard"
ws1['A1'].font = Font(bold=True, size=16)
ws1['A2'] = f"Prepared: {datetime.now().strftime('%B %d, %Y')}"
ws1['A3'] = "Owner: Tarik Perkins | Platform: Airbnb"
ws1['A4'] = "Properties: Upstate New York (Morabeza, Sunset Studio)"

# Monthly pivot
months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
us_properties = ['Morabeza (Studio)', 'Sunset Studio']

row = 6
ws1.cell(row=row, column=1, value="Property / Month")
for i, m in enumerate(months):
    ws1.cell(row=row, column=i+2, value=m)
ws1.cell(row=row, column=14, value="TOTAL")
style_header_row(ws1, row, 14)

for prop in us_properties:
    row += 1
    ws1.cell(row=row, column=1, value=prop)
    ws1.cell(row=row, column=1).font = Font(bold=True)
    prop_total = 0
    for m_idx in range(1, 13):
        month_income = sum(r['host_payout']['you_earn'] for r in us_bookings_2025 
                         if r['_norm_property'] == prop and get_month(r['check_in_date']) == m_idx)
        ws1.cell(row=row, column=m_idx+1, value=month_income if month_income > 0 else None)
        ws1.cell(row=row, column=m_idx+1).number_format = money_format
        prop_total += month_income
    ws1.cell(row=row, column=14, value=prop_total)
    ws1.cell(row=row, column=14).number_format = money_format
    ws1.cell(row=row, column=14).font = Font(bold=True)

# Total row
row += 1
ws1.cell(row=row, column=1, value="TOTAL US HOST EARNINGS")
ws1.cell(row=row, column=1).font = Font(bold=True, size=11)
grand_earn = 0
for m_idx in range(1, 13):
    month_all = sum(r['host_payout']['you_earn'] for r in us_bookings_2025 if get_month(r['check_in_date']) == m_idx)
    ws1.cell(row=row, column=m_idx+1, value=month_all if month_all > 0 else None)
    ws1.cell(row=row, column=m_idx+1).number_format = money_format
    ws1.cell(row=row, column=m_idx+1).font = Font(bold=True)
    grand_earn += month_all
ws1.cell(row=row, column=14, value=grand_earn)
ws1.cell(row=row, column=14).number_format = money_format
ws1.cell(row=row, column=14).font = Font(bold=True, size=12)

# Summary stats
row += 3
ws1.cell(row=row, column=1, value="US Property Summary Statistics")
ws1.cell(row=row, column=1).font = Font(bold=True, size=14)
row += 1
stat_headers = ['Metric', 'Morabeza (Studio)', 'Sunset Studio', 'Total US']
for i, h in enumerate(stat_headers):
    ws1.cell(row=row, column=i+1, value=h)
style_header_row(ws1, row, 4)

metrics = [
    ('Bookings', lambda b: 1),
    ('Total Nights', lambda b: b.get('num_nights', 0)),
    ('Gross Revenue (Guest Paid)', lambda b: b['guest_paid']['total_usd']),
    ('Cleaning Fees Collected', lambda b: b['guest_paid']['cleaning_fee']),
    ('Occupancy Taxes Collected', lambda b: b['guest_paid']['occupancy_taxes']),
    ('Platform Commission', lambda b: abs(b['host_payout']['host_service_fee'])),
    ('Net Host Earnings (You Earn)', lambda b: b['host_payout']['you_earn']),
    ('Avg Nightly Rate', None),
    ('Avg Booking Value', None),
]

for metric_name, calc_fn in metrics:
    if calc_fn is None:
        continue
    row += 1
    ws1.cell(row=row, column=1, value=metric_name)
    for col_idx, prop in enumerate(us_properties, 2):
        prop_bookings = [b for b in us_bookings_2025 if b['_norm_property'] == prop]
        val = sum(calc_fn(b) for b in prop_bookings)
        ws1.cell(row=row, column=col_idx, value=val)
        if 'Revenue' in metric_name or 'Fee' in metric_name or 'Tax' in metric_name or 'Commission' in metric_name or 'Earn' in metric_name:
            ws1.cell(row=row, column=col_idx).number_format = money_format
    # Total
    all_val = sum(calc_fn(b) for b in us_bookings_2025)
    ws1.cell(row=row, column=4, value=all_val)
    if 'Revenue' in metric_name or 'Fee' in metric_name or 'Tax' in metric_name or 'Commission' in metric_name or 'Earn' in metric_name:
        ws1.cell(row=row, column=4).number_format = money_format
    ws1.cell(row=row, column=4).font = Font(bold=True)

# Avg nightly rate
row += 1
ws1.cell(row=row, column=1, value="Avg Nightly Rate")
for col_idx, prop in enumerate(us_properties, 2):
    prop_bookings = [b for b in us_bookings_2025 if b['_norm_property'] == prop]
    total_nights = sum(b.get('num_nights', 0) for b in prop_bookings)
    total_room = sum(b['guest_paid']['room_subtotal'] for b in prop_bookings)
    ws1.cell(row=row, column=col_idx, value=total_room / total_nights if total_nights > 0 else 0)
    ws1.cell(row=row, column=col_idx).number_format = money_format

# Avg booking value
row += 1
ws1.cell(row=row, column=1, value="Avg Booking Value (Host Earn)")
for col_idx, prop in enumerate(us_properties, 2):
    prop_bookings = [b for b in us_bookings_2025 if b['_norm_property'] == prop]
    count = len(prop_bookings)
    total_earn = sum(b['host_payout']['you_earn'] for b in prop_bookings)
    ws1.cell(row=row, column=col_idx, value=total_earn / count if count > 0 else 0)
    ws1.cell(row=row, column=col_idx).number_format = money_format

auto_width(ws1)

# ============================================================
# SHEET 2: US Booking Details
# ============================================================
ws2 = wb.create_sheet("US Booking Details")

ws2['A1'] = "2025 US Property Booking Details"
ws2['A1'].font = Font(bold=True, size=16)
ws2['A2'] = "Complete financial breakdown per reservation"

row = 4
detail_headers = [
    'Check-In', 'Check-Out', 'Nights', 'Property', 'Guest Name', 'Guests',
    'Conf. Code', 'Nightly Rate', 'Room Subtotal', 'Cleaning Fee',
    'Guest Service Fee', 'Occupancy Tax', 'Total Guest Paid',
    'Host Room Fee', 'Host Cleaning', 'Rate Adjustment', 
    'Host Service Fee', 'Fee %', 'Taxes Passed', 'Host Earns (Net)'
]
for i, h in enumerate(detail_headers):
    ws2.cell(row=row, column=i+1, value=h)
style_header_row(ws2, row, len(detail_headers))

for b in sorted(us_bookings_2025, key=lambda x: x.get('check_in_date', '')):
    row += 1
    ws2.cell(row=row, column=1, value=b['check_in_date'])
    ws2.cell(row=row, column=2, value=b['check_out_date'])
    ws2.cell(row=row, column=3, value=b.get('num_nights', 0))
    ws2.cell(row=row, column=4, value=b['_norm_property'])
    ws2.cell(row=row, column=5, value=b['guest_name'])
    ws2.cell(row=row, column=6, value=b.get('guest_count', 0))
    ws2.cell(row=row, column=7, value=b['confirmation_code'])
    ws2.cell(row=row, column=8, value=b['guest_paid']['nightly_rate'])
    ws2.cell(row=row, column=8).number_format = money_format
    ws2.cell(row=row, column=9, value=b['guest_paid']['room_subtotal'])
    ws2.cell(row=row, column=9).number_format = money_format
    ws2.cell(row=row, column=10, value=b['guest_paid']['cleaning_fee'])
    ws2.cell(row=row, column=10).number_format = money_format
    ws2.cell(row=row, column=11, value=b['guest_paid']['guest_service_fee'])
    ws2.cell(row=row, column=11).number_format = money_format
    ws2.cell(row=row, column=12, value=b['guest_paid']['occupancy_taxes'])
    ws2.cell(row=row, column=12).number_format = money_format
    ws2.cell(row=row, column=13, value=b['guest_paid']['total_usd'])
    ws2.cell(row=row, column=13).number_format = money_format
    ws2.cell(row=row, column=13).font = Font(bold=True)
    ws2.cell(row=row, column=14, value=b['host_payout']['room_fee'])
    ws2.cell(row=row, column=14).number_format = money_format
    ws2.cell(row=row, column=15, value=b['host_payout']['cleaning_fee'])
    ws2.cell(row=row, column=15).number_format = money_format
    ws2.cell(row=row, column=16, value=b['host_payout']['nightly_rate_adjustment'])
    ws2.cell(row=row, column=16).number_format = money_format
    ws2.cell(row=row, column=17, value=b['host_payout']['host_service_fee'])
    ws2.cell(row=row, column=17).number_format = money_format
    ws2.cell(row=row, column=18, value=b['host_payout']['host_service_fee_percent'])
    ws2.cell(row=row, column=18).number_format = '0.0'
    ws2.cell(row=row, column=19, value=b['host_payout']['property_use_taxes'])
    ws2.cell(row=row, column=19).number_format = money_format
    ws2.cell(row=row, column=20, value=b['host_payout']['you_earn'])
    ws2.cell(row=row, column=20).number_format = money_format
    ws2.cell(row=row, column=20).font = Font(bold=True)

# Totals row
row += 1
ws2.cell(row=row, column=1, value="TOTALS")
ws2.cell(row=row, column=1).font = Font(bold=True)
ws2.cell(row=row, column=3, value=sum(b.get('num_nights', 0) for b in us_bookings_2025))
ws2.cell(row=row, column=3).font = Font(bold=True)
ws2.cell(row=row, column=9, value=sum(b['guest_paid']['room_subtotal'] for b in us_bookings_2025))
ws2.cell(row=row, column=9).number_format = money_format
ws2.cell(row=row, column=9).font = Font(bold=True)
ws2.cell(row=row, column=10, value=sum(b['guest_paid']['cleaning_fee'] for b in us_bookings_2025))
ws2.cell(row=row, column=10).number_format = money_format
ws2.cell(row=row, column=10).font = Font(bold=True)
ws2.cell(row=row, column=11, value=sum(b['guest_paid']['guest_service_fee'] for b in us_bookings_2025))
ws2.cell(row=row, column=11).number_format = money_format
ws2.cell(row=row, column=11).font = Font(bold=True)
ws2.cell(row=row, column=12, value=sum(b['guest_paid']['occupancy_taxes'] for b in us_bookings_2025))
ws2.cell(row=row, column=12).number_format = money_format
ws2.cell(row=row, column=12).font = Font(bold=True)
ws2.cell(row=row, column=13, value=sum(b['guest_paid']['total_usd'] for b in us_bookings_2025))
ws2.cell(row=row, column=13).number_format = money_format
ws2.cell(row=row, column=13).font = Font(bold=True)
ws2.cell(row=row, column=17, value=sum(b['host_payout']['host_service_fee'] for b in us_bookings_2025))
ws2.cell(row=row, column=17).number_format = money_format
ws2.cell(row=row, column=17).font = Font(bold=True)
ws2.cell(row=row, column=20, value=sum(b['host_payout']['you_earn'] for b in us_bookings_2025))
ws2.cell(row=row, column=20).number_format = money_format
ws2.cell(row=row, column=20).font = Font(bold=True, size=12)

auto_width(ws2)

# ============================================================
# SHEET 3: US Expenses
# ============================================================
ws3 = wb.create_sheet("US Expenses")

ws3['A1'] = "2025 US Property Expenses"
ws3['A1'].font = Font(bold=True, size=16)
ws3['A2'] = "Sources: Amazon, Walmart (from tariqp@gmail.com)"
ws3['A3'] = "Note: Only property-related expenses included. Personal purchases excluded."

row = 5
expense_headers = ['Date', 'Vendor', 'Category', 'Description', 'Amount', 'Property', 'Source']
for i, h in enumerate(expense_headers):
    ws3.cell(row=row, column=i+1, value=h)
style_header_row(ws3, row, len(expense_headers))

sorted_expenses = sorted(property_expenses, key=lambda x: x.get('date', ''))
total_expenses = 0

for e in sorted_expenses:
    row += 1
    ws3.cell(row=row, column=1, value=e.get('date', ''))
    ws3.cell(row=row, column=2, value=e.get('vendor', ''))
    ws3.cell(row=row, column=3, value=e.get('category', ''))
    ws3.cell(row=row, column=4, value=e.get('items_description', '')[:80])
    amt = e.get('total_amount', 0)
    ws3.cell(row=row, column=5, value=amt)
    ws3.cell(row=row, column=5).number_format = money_format
    ws3.cell(row=row, column=6, value=e.get('likely_property', 'Unknown'))
    ws3.cell(row=row, column=7, value=e.get('_source', ''))
    total_expenses += amt

row += 2
ws3.cell(row=row, column=4, value="TOTAL PROPERTY EXPENSES")
ws3.cell(row=row, column=4).font = Font(bold=True)
ws3.cell(row=row, column=5, value=total_expenses)
ws3.cell(row=row, column=5).number_format = money_format
ws3.cell(row=row, column=5).font = Font(bold=True)

# Category summary
row += 3
ws3.cell(row=row, column=1, value="Expense Summary by Category")
ws3.cell(row=row, column=1).font = Font(bold=True, size=14)
row += 1
cat_headers = ['Category', 'Count', 'Total Amount']
for i, h in enumerate(cat_headers):
    ws3.cell(row=row, column=i+1, value=h)
style_header_row(ws3, row, 3)

by_cat = {}
for e in sorted_expenses:
    cat = e.get('category', 'other')
    if cat not in by_cat:
        by_cat[cat] = {'count': 0, 'total': 0}
    by_cat[cat]['count'] += 1
    by_cat[cat]['total'] += e.get('total_amount', 0)

for cat, data in sorted(by_cat.items(), key=lambda x: -x[1]['total']):
    row += 1
    ws3.cell(row=row, column=1, value=cat.title())
    ws3.cell(row=row, column=2, value=data['count'])
    ws3.cell(row=row, column=3, value=data['total'])
    ws3.cell(row=row, column=3).number_format = money_format

auto_width(ws3)

# ============================================================
# SHEET 4: US P&L Summary
# ============================================================
ws4 = wb.create_sheet("US P&L Summary")

ws4['A1'] = "2025 US Property Profit & Loss Summary"
ws4['A1'].font = Font(bold=True, size=16)
ws4['A2'] = "Schedule E Ready — Short-Term Rental Properties"
ws4['A3'] = f"Prepared: {datetime.now().strftime('%B %d, %Y')}"

row = 5
pnl_headers = ['Line Item', 'Morabeza (Studio)', 'Sunset Studio', 'TOTAL US']
for i, h in enumerate(pnl_headers):
    ws4.cell(row=row, column=i+1, value=h)
style_header_row(ws4, row, len(pnl_headers))

# Calculate per property
pnl_data = {}
for prop in us_properties:
    prop_bookings = [b for b in us_bookings_2025 if b['_norm_property'] == prop]
    pnl_data[prop] = {
        'gross_revenue': sum(b['guest_paid']['total_usd'] for b in prop_bookings),
        'host_earnings': sum(b['host_payout']['you_earn'] for b in prop_bookings),
        'cleaning_fees': sum(b['guest_paid']['cleaning_fee'] for b in prop_bookings),
        'occupancy_taxes': sum(b['guest_paid']['occupancy_taxes'] for b in prop_bookings),
        'platform_fees': sum(abs(b['host_payout']['host_service_fee']) for b in prop_bookings),
        'bookings': len(prop_bookings),
        'nights': sum(b.get('num_nights', 0) for b in prop_bookings)
    }

# Allocate expenses
general_exp = sum(e.get('total_amount', 0) for e in sorted_expenses 
                 if e.get('likely_property', '') in ['General (all properties)', 'General', 'Unknown', ''])
per_prop_share = general_exp / len(us_properties) if us_properties else 0

for prop in us_properties:
    prop_specific = sum(e.get('total_amount', 0) for e in sorted_expenses 
                       if normalize_property(e.get('likely_property', '')) == prop)
    pnl_data[prop]['expenses'] = prop_specific + per_prop_share

# P&L rows
pnl_lines = [
    ('INCOME', None, True),
    ('Gross Revenue (Total Guest Paid)', 'gross_revenue', False),
    ('Less: Occupancy Taxes (pass-through)', 'occupancy_taxes', False),
    ('Less: Airbnb Platform Commission', 'platform_fees', False),
    ('Net Host Earnings', 'host_earnings', True),
    ('', None, False),
    ('EXPENSES', None, True),
    ('Property Expenses (supplies, maintenance, etc.)', 'expenses', False),
    ('', None, False),
    ('NET PROFIT / (LOSS)', None, True),
]

for line_name, key, is_bold in pnl_lines:
    row += 1
    ws4.cell(row=row, column=1, value=line_name)
    if is_bold:
        ws4.cell(row=row, column=1).font = Font(bold=True)
    
    if key:
        for col_idx, prop in enumerate(us_properties, 2):
            val = pnl_data[prop].get(key, 0)
            ws4.cell(row=row, column=col_idx, value=val)
            ws4.cell(row=row, column=col_idx).number_format = money_format
            if is_bold:
                ws4.cell(row=row, column=col_idx).font = Font(bold=True)
        # Total
        total_val = sum(pnl_data[p].get(key, 0) for p in us_properties)
        ws4.cell(row=row, column=4, value=total_val)
        ws4.cell(row=row, column=4).number_format = money_format
        ws4.cell(row=row, column=4).font = Font(bold=True)
    elif line_name == 'NET PROFIT / (LOSS)':
        for col_idx, prop in enumerate(us_properties, 2):
            net = pnl_data[prop]['host_earnings'] - pnl_data[prop]['expenses']
            ws4.cell(row=row, column=col_idx, value=net)
            ws4.cell(row=row, column=col_idx).number_format = money_format
            ws4.cell(row=row, column=col_idx).font = Font(bold=True, color="006400" if net >= 0 else "8B0000")
        total_net = sum(pnl_data[p]['host_earnings'] - pnl_data[p]['expenses'] for p in us_properties)
        ws4.cell(row=row, column=4, value=total_net)
        ws4.cell(row=row, column=4).number_format = money_format
        ws4.cell(row=row, column=4).font = Font(bold=True, size=12, color="006400" if total_net >= 0 else "8B0000")

# Notes
row += 3
ws4.cell(row=row, column=1, value="Notes for CPA:")
ws4.cell(row=row, column=1).font = Font(bold=True, size=12)
notes = [
    "1. Income data from Airbnb reservation confirmation emails (tarikp.us@gmail.com).",
    "2. 'Net Host Earnings' = what Airbnb deposits after their commission and before expenses.",
    "3. Occupancy taxes are collected by Airbnb and remitted to local authorities — may not be taxable income.",
    "4. Expenses from Amazon/Walmart purchase emails. General expenses split evenly across US properties.",
    "5. Additional expenses (utilities, insurance, mortgage interest, depreciation) not captured here.",
    "6. Vrbo bookings exist in calendar but financial data not available from email scrape.",
]
for note in notes:
    row += 1
    ws4.cell(row=row, column=1, value=note)

auto_width(ws4)

# ============================================================
# SHEET 5: Jamaica Income
# ============================================================
ws5 = wb.create_sheet("Jamaica Income")

ws5['A1'] = "2025 Jamaica Property Income"
ws5['A1'].font = Font(bold=True, size=16)
ws5['A2'] = "Property: The Artiste's Boutique (Bohemian Lodge)"
ws5['A3'] = "Location: Jamaica | Platform: Airbnb"

# Monthly summary
row = 5
ws5.cell(row=row, column=1, value="Monthly Summary")
ws5.cell(row=row, column=1).font = Font(bold=True, size=14)
row += 1
m_headers = ['Month', 'Bookings', 'Nights', 'Gross Revenue', 'Platform Fee', 'Host Earns']
for i, h in enumerate(m_headers):
    ws5.cell(row=row, column=i+1, value=h)
style_header_row(ws5, row, len(m_headers))

total_j_earn = 0
total_j_gross = 0
total_j_fee = 0
total_j_nights = 0
total_j_bookings = 0

for m_idx in range(1, 13):
    m_bookings = [b for b in jamaica_bookings_2025 if get_month(b['check_in_date']) == m_idx]
    if not m_bookings:
        continue
    row += 1
    m_earn = sum(b['host_payout']['you_earn'] for b in m_bookings)
    m_gross = sum(b['guest_paid']['total_usd'] for b in m_bookings)
    m_fee = sum(abs(b['host_payout']['host_service_fee']) for b in m_bookings)
    m_nights = sum(b.get('num_nights', 0) for b in m_bookings)
    
    ws5.cell(row=row, column=1, value=months[m_idx-1])
    ws5.cell(row=row, column=2, value=len(m_bookings))
    ws5.cell(row=row, column=3, value=m_nights)
    ws5.cell(row=row, column=4, value=m_gross)
    ws5.cell(row=row, column=4).number_format = money_format
    ws5.cell(row=row, column=5, value=m_fee)
    ws5.cell(row=row, column=5).number_format = money_format
    ws5.cell(row=row, column=6, value=m_earn)
    ws5.cell(row=row, column=6).number_format = money_format
    
    total_j_earn += m_earn
    total_j_gross += m_gross
    total_j_fee += m_fee
    total_j_nights += m_nights
    total_j_bookings += len(m_bookings)

# Total
row += 1
ws5.cell(row=row, column=1, value="TOTAL 2025")
ws5.cell(row=row, column=1).font = Font(bold=True)
ws5.cell(row=row, column=2, value=total_j_bookings)
ws5.cell(row=row, column=2).font = Font(bold=True)
ws5.cell(row=row, column=3, value=total_j_nights)
ws5.cell(row=row, column=3).font = Font(bold=True)
ws5.cell(row=row, column=4, value=total_j_gross)
ws5.cell(row=row, column=4).number_format = money_format
ws5.cell(row=row, column=4).font = Font(bold=True)
ws5.cell(row=row, column=5, value=total_j_fee)
ws5.cell(row=row, column=5).number_format = money_format
ws5.cell(row=row, column=5).font = Font(bold=True)
ws5.cell(row=row, column=6, value=total_j_earn)
ws5.cell(row=row, column=6).number_format = money_format
ws5.cell(row=row, column=6).font = Font(bold=True, size=12)

auto_width(ws5)

# ============================================================
# SHEET 6: Jamaica Booking Details
# ============================================================
ws6 = wb.create_sheet("Jamaica Booking Details")

ws6['A1'] = "Jamaica Booking Details — The Artiste's Boutique"
ws6['A1'].font = Font(bold=True, size=16)
ws6['A2'] = "2025 bookings + confirmed future bookings"

row = 4
j_headers = [
    'Check-In', 'Check-Out', 'Nights', 'Guest Name', 'Guests',
    'Conf. Code', 'Nightly Rate', 'Room Subtotal', 'Cleaning Fee',
    'Guest Service Fee', 'Occupancy Tax', 'Total Guest Paid',
    'Host Service Fee', 'Host Earns (Net)', 'Year'
]
for i, h in enumerate(j_headers):
    ws6.cell(row=row, column=i+1, value=h)
style_header_row(ws6, row, len(j_headers))

all_jamaica = jamaica_bookings_2025 + jamaica_bookings_future
for b in sorted(all_jamaica, key=lambda x: x.get('check_in_date', '')):
    row += 1
    year = b['check_in_date'][:4] if b.get('check_in_date') else ''
    ws6.cell(row=row, column=1, value=b['check_in_date'])
    ws6.cell(row=row, column=2, value=b['check_out_date'])
    ws6.cell(row=row, column=3, value=b.get('num_nights', 0))
    ws6.cell(row=row, column=4, value=b['guest_name'])
    ws6.cell(row=row, column=5, value=b.get('guest_count', 0))
    ws6.cell(row=row, column=6, value=b['confirmation_code'])
    ws6.cell(row=row, column=7, value=b['guest_paid']['nightly_rate'])
    ws6.cell(row=row, column=7).number_format = money_format
    ws6.cell(row=row, column=8, value=b['guest_paid']['room_subtotal'])
    ws6.cell(row=row, column=8).number_format = money_format
    ws6.cell(row=row, column=9, value=b['guest_paid']['cleaning_fee'])
    ws6.cell(row=row, column=9).number_format = money_format
    ws6.cell(row=row, column=10, value=b['guest_paid']['guest_service_fee'])
    ws6.cell(row=row, column=10).number_format = money_format
    ws6.cell(row=row, column=11, value=b['guest_paid']['occupancy_taxes'])
    ws6.cell(row=row, column=11).number_format = money_format
    ws6.cell(row=row, column=12, value=b['guest_paid']['total_usd'])
    ws6.cell(row=row, column=12).number_format = money_format
    ws6.cell(row=row, column=12).font = Font(bold=True)
    ws6.cell(row=row, column=13, value=b['host_payout']['host_service_fee'])
    ws6.cell(row=row, column=13).number_format = money_format
    ws6.cell(row=row, column=14, value=b['host_payout']['you_earn'])
    ws6.cell(row=row, column=14).number_format = money_format
    ws6.cell(row=row, column=14).font = Font(bold=True)
    ws6.cell(row=row, column=15, value=year)

# Totals
row += 1
ws6.cell(row=row, column=1, value="2025 TOTALS")
ws6.cell(row=row, column=1).font = Font(bold=True)
ws6.cell(row=row, column=3, value=sum(b.get('num_nights', 0) for b in jamaica_bookings_2025))
ws6.cell(row=row, column=3).font = Font(bold=True)
ws6.cell(row=row, column=12, value=sum(b['guest_paid']['total_usd'] for b in jamaica_bookings_2025))
ws6.cell(row=row, column=12).number_format = money_format
ws6.cell(row=row, column=12).font = Font(bold=True)
ws6.cell(row=row, column=14, value=sum(b['host_payout']['you_earn'] for b in jamaica_bookings_2025))
ws6.cell(row=row, column=14).number_format = money_format
ws6.cell(row=row, column=14).font = Font(bold=True, size=12)

auto_width(ws6)

# ============================================================
# SAVE WORKBOOK
# ============================================================
output_path = '/home/ubuntu/tax_prep/Tarik_Perkins_2025_Property_Tax_Workbook_v2.xlsx'
wb.save(output_path)

# Final summary
us_total_earn = sum(b['host_payout']['you_earn'] for b in us_bookings_2025)
us_total_gross = sum(b['guest_paid']['total_usd'] for b in us_bookings_2025)
us_total_fee = sum(abs(b['host_payout']['host_service_fee']) for b in us_bookings_2025)
j_total_earn = sum(b['host_payout']['you_earn'] for b in jamaica_bookings_2025)
j_total_gross = sum(b['guest_paid']['total_usd'] for b in jamaica_bookings_2025)

print(f"\n{'='*60}")
print(f"2025 PROPERTY TAX WORKBOOK v2 — COMPLETE")
print(f"{'='*60}")
print(f"\nUS Properties (Morabeza + Sunset Studio):")
print(f"  Bookings: {len(us_bookings_2025)} | Nights: {sum(b.get('num_nights',0) for b in us_bookings_2025)}")
print(f"  Gross Revenue (Guest Paid): ${us_total_gross:,.2f}")
print(f"  Platform Commission:         ${us_total_fee:,.2f}")
print(f"  Net Host Earnings:           ${us_total_earn:,.2f}")
print(f"  Property Expenses:           ${total_expenses:,.2f}")
print(f"  NET PROFIT:                  ${us_total_earn - total_expenses:,.2f}")
print(f"\nJamaica (The Artiste's Boutique):")
print(f"  Bookings: {len(jamaica_bookings_2025)} | Nights: {sum(b.get('num_nights',0) for b in jamaica_bookings_2025)}")
print(f"  Gross Revenue (Guest Paid): ${j_total_gross:,.2f}")
print(f"  Net Host Earnings:           ${j_total_earn:,.2f}")
print(f"\nFuture Bookings (2026):")
print(f"  Jamaica: {len(jamaica_bookings_future)} bookings")
print(f"  Earning: ${sum(b['host_payout']['you_earn'] for b in jamaica_bookings_future):,.2f}")
print(f"\nSheets: US Income Dashboard, US Booking Details, US Expenses,")
print(f"        US P&L Summary, Jamaica Income, Jamaica Booking Details")
print(f"\nSaved to: {output_path}")
