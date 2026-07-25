#!/usr/bin/env python3
"""Add CBNA Checking, CBNA Savings, and Booking.com tabs to the main workbook."""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

# Load the workbook
wb = openpyxl.load_workbook('Tarik_Perkins_2025_Property_Tax_Workbook_v4.xlsx')

# Styling
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
credit_font = Font(color="006100")
debit_font = Font(color="9C0006")
credit_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
debit_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# ============================================================
# TAB 1: CBNA Checking
# ============================================================
print("Adding CBNA Checking tab...")
ws_cbna = wb.create_sheet("CBNA Checking (x8792)")

# Read the CSV
cbna_df = pd.read_csv('cbna_checking_transactions.csv')

# Title row
ws_cbna['A1'] = 'CBNA Checking Account (x8792) - Carefree Checking'
ws_cbna['A1'].font = Font(bold=True, size=14)
ws_cbna['A2'] = 'Owoicho E Ameh / Tarik J Perkins - 4693 State Route 414, Burdett NY'
ws_cbna['A2'].font = Font(italic=True, size=10)
ws_cbna['A3'] = 'Statement Period: Jul 2024 - Jun 2026 | NO STR/VRBO deposits found in this account'
ws_cbna['A3'].font = Font(bold=True, size=10, color="FF0000")

# Headers
headers = ['Date', 'Description', 'Amount', 'Balance', 'Type']
for col, header in enumerate(headers, 1):
    cell = ws_cbna.cell(row=5, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

# Data
for idx, row in cbna_df.iterrows():
    r = idx + 6
    ws_cbna.cell(row=r, column=1, value=row['Date']).border = thin_border
    ws_cbna.cell(row=r, column=2, value=row['Description']).border = thin_border
    
    amt_cell = ws_cbna.cell(row=r, column=3, value=row['Amount'])
    amt_cell.border = thin_border
    amt_cell.number_format = '#,##0.00'
    
    if pd.notna(row['Balance']):
        bal_cell = ws_cbna.cell(row=r, column=4, value=row['Balance'])
        bal_cell.border = thin_border
        bal_cell.number_format = '#,##0.00'
    
    type_cell = ws_cbna.cell(row=r, column=5, value=row['Type'])
    type_cell.border = thin_border
    
    # Color code credits/debits
    if row['Type'] == 'Credit':
        amt_cell.font = credit_font
        amt_cell.fill = credit_fill
    else:
        amt_cell.font = debit_font
        amt_cell.fill = debit_fill

# Column widths
ws_cbna.column_dimensions['A'].width = 12
ws_cbna.column_dimensions['B'].width = 65
ws_cbna.column_dimensions['C'].width = 14
ws_cbna.column_dimensions['D'].width = 14
ws_cbna.column_dimensions['E'].width = 10

# Summary at bottom
summary_row = len(cbna_df) + 7
ws_cbna.cell(row=summary_row, column=1, value='SUMMARY').font = Font(bold=True, size=11)
ws_cbna.cell(row=summary_row+1, column=1, value='Total Credits:')
ws_cbna.cell(row=summary_row+1, column=3, value=cbna_df[cbna_df['Type']=='Credit']['Amount'].sum())
ws_cbna.cell(row=summary_row+1, column=3).number_format = '$#,##0.00'
ws_cbna.cell(row=summary_row+2, column=1, value='Total Debits:')
ws_cbna.cell(row=summary_row+2, column=3, value=cbna_df[cbna_df['Type']=='Debit']['Amount'].sum())
ws_cbna.cell(row=summary_row+2, column=3).number_format = '$#,##0.00'
ws_cbna.cell(row=summary_row+4, column=1, value='NOTE: QBO shows 2 VRBO deposits to "CBNA Checking" in Jan-Feb 2024 ($607.20 + $883.20)')
ws_cbna.cell(row=summary_row+4, column=1).font = Font(bold=True, color="FF0000")
ws_cbna.cell(row=summary_row+5, column=1, value='These statements start Jul 2024. Need Jan-Jun 2024 statements to verify those deposits.')
ws_cbna.cell(row=summary_row+5, column=1).font = Font(italic=True)

print(f"  Added {len(cbna_df)} transactions")

# ============================================================
# TAB 2: CBNA Savings
# ============================================================
print("Adding CBNA Savings tab...")
ws_sav = wb.create_sheet("CBNA Savings (x8774)")

cbna_sav_df = pd.read_csv('cbna_savings_transactions.csv')

ws_sav['A1'] = 'CBNA Savings Account (x8774) - Free Savings'
ws_sav['A1'].font = Font(bold=True, size=14)
ws_sav['A2'] = 'Minimal activity account - small transfers only'
ws_sav['A2'].font = Font(italic=True, size=10)

# Headers
for col, header in enumerate(headers, 1):
    cell = ws_sav.cell(row=4, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border

# Data
for idx, row in cbna_sav_df.iterrows():
    r = idx + 5
    ws_sav.cell(row=r, column=1, value=row['Date']).border = thin_border
    ws_sav.cell(row=r, column=2, value=row['Description']).border = thin_border
    ws_sav.cell(row=r, column=3, value=row['Amount']).border = thin_border
    if pd.notna(row['Balance']):
        ws_sav.cell(row=r, column=4, value=row['Balance']).border = thin_border
    ws_sav.cell(row=r, column=5, value=row['Type']).border = thin_border

ws_sav.column_dimensions['A'].width = 12
ws_sav.column_dimensions['B'].width = 50
ws_sav.column_dimensions['C'].width = 14
ws_sav.column_dimensions['D'].width = 14
ws_sav.column_dimensions['E'].width = 10

print(f"  Added {len(cbna_sav_df)} transactions")

# ============================================================
# TAB 3: Booking.com Reservations
# ============================================================
print("Adding Booking.com tab...")
ws_bk = wb.create_sheet("Booking.com")

# Read the XLS
bk_df = pd.read_excel('gdrive_data/booking_com/Reservations_2024-01-01_2026-12-31.xls')
bk_df['arrival_dt'] = pd.to_datetime(bk_df['Arrival'], format='mixed')

ws_bk['A1'] = 'Booking.com Reservation Data (2024-2026)'
ws_bk['A1'].font = Font(bold=True, size=14)
ws_bk['A2'] = 'Source: admin.booking.com extranet export'
ws_bk['A2'].font = Font(italic=True, size=10)

# Summary section
ws_bk['A4'] = 'SUMMARY BY YEAR AND PROPERTY (OK Status Only)'
ws_bk['A4'].font = Font(bold=True, size=11)

summary_row = 5
for year in sorted(bk_df['arrival_dt'].dt.year.unique()):
    yr_df = bk_df[(bk_df['arrival_dt'].dt.year == year) & (bk_df['Status'] == 'OK')]
    ws_bk.cell(row=summary_row, column=1, value=f'{year}:').font = Font(bold=True)
    ws_bk.cell(row=summary_row, column=2, value=f'{len(yr_df)} confirmed bookings')
    ws_bk.cell(row=summary_row, column=3, value=yr_df['Total Payment'].sum())
    ws_bk.cell(row=summary_row, column=3).number_format = '$#,##0.00'
    summary_row += 1
    for prop in yr_df['Property Name'].unique():
        prop_df = yr_df[yr_df['Property Name'] == prop]
        ws_bk.cell(row=summary_row, column=2, value=f'  {prop}')
        ws_bk.cell(row=summary_row, column=3, value=prop_df['Total Payment'].sum())
        ws_bk.cell(row=summary_row, column=3).number_format = '$#,##0.00'
        ws_bk.cell(row=summary_row, column=4, value=f'{len(prop_df)} bookings')
        summary_row += 1

# Full data table
data_start = summary_row + 2
ws_bk.cell(row=data_start, column=1, value='ALL RESERVATIONS').font = Font(bold=True, size=11)

bk_headers = ['Property Name', 'Booker Name', 'Arrival', 'Departure', 'Booked on', 
              'Status', 'Total Payment', 'Commission', 'Reservation Number']
for col, header in enumerate(bk_headers, 1):
    cell = ws_bk.cell(row=data_start+1, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border

# Sort by arrival date
bk_df_sorted = bk_df.sort_values('arrival_dt')

ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
cancel_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
noshow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

for idx, (_, row) in enumerate(bk_df_sorted.iterrows()):
    r = data_start + 2 + idx
    ws_bk.cell(row=r, column=1, value=row['Property Name']).border = thin_border
    ws_bk.cell(row=r, column=2, value=row['Booker Name']).border = thin_border
    ws_bk.cell(row=r, column=3, value=row['Arrival']).border = thin_border
    ws_bk.cell(row=r, column=4, value=row['Departure']).border = thin_border
    ws_bk.cell(row=r, column=5, value=row['Booked on']).border = thin_border
    
    status_cell = ws_bk.cell(row=r, column=6, value=row['Status'])
    status_cell.border = thin_border
    if row['Status'] == 'OK':
        status_cell.fill = ok_fill
    elif row['Status'] == 'Canceled':
        status_cell.fill = cancel_fill
    elif row['Status'] == 'No-show':
        status_cell.fill = noshow_fill
    
    ws_bk.cell(row=r, column=7, value=row['Total Payment']).border = thin_border
    ws_bk.cell(row=r, column=7).number_format = '$#,##0.00'
    
    if pd.notna(row['Commission']):
        ws_bk.cell(row=r, column=8, value=row['Commission']).border = thin_border
        ws_bk.cell(row=r, column=8).number_format = '$#,##0.00'
    
    ws_bk.cell(row=r, column=9, value=row['Reservation Number']).border = thin_border

# Column widths
ws_bk.column_dimensions['A'].width = 38
ws_bk.column_dimensions['B'].width = 28
ws_bk.column_dimensions['C'].width = 20
ws_bk.column_dimensions['D'].width = 20
ws_bk.column_dimensions['E'].width = 20
ws_bk.column_dimensions['F'].width = 12
ws_bk.column_dimensions['G'].width = 14
ws_bk.column_dimensions['H'].width = 14
ws_bk.column_dimensions['I'].width = 18

print(f"  Added {len(bk_df)} reservations")

# ============================================================
# TAB 4: PayPal (placeholder with note)
# ============================================================
print("Adding PayPal tab (placeholder)...")
ws_pp = wb.create_sheet("PayPal")
ws_pp['A1'] = 'PayPal Transaction Data'
ws_pp['A1'].font = Font(bold=True, size=14)
ws_pp['A3'] = 'STATUS: Data file is empty (headers only, no transaction rows)'
ws_pp['A3'].font = Font(bold=True, color="FF0000")
ws_pp['A4'] = 'The uploaded PayPal CSV contains only column headers with no data.'
ws_pp['A5'] = 'Please re-export from PayPal with the date range 01/01/2024 - 12/31/2024.'
ws_pp['A7'] = 'QBO shows 2 PayPal deposits in 2024:'
ws_pp['A7'].font = Font(bold=True)
ws_pp['A8'] = '  - $359.00 (date TBD)'
ws_pp['A9'] = '  - $1,279.00 (date TBD)'
ws_pp['A10'] = '  Total: $1,638.00 to Bank of America Savings 0108'

# ============================================================
# Save
# ============================================================
output_file = 'Tarik_Perkins_2025_Property_Tax_Workbook_v5.xlsx'
wb.save(output_file)
print(f"\nWorkbook saved as: {output_file}")
print(f"Sheets: {wb.sheetnames}")
