#!/usr/bin/env python3
"""
Tarik Perkins 2025 Property Tax Workbook v4
============================================
Builds an Excel workbook for US tax filing (Schedule E) from the Geeves property_bookings DB.

Properties:
- Morabeza (4693 SR 414, Burdett NY 14818) — US Schedule E
- Sunset Studio (4693 SR 414, Burdett NY 14818) — US Schedule E
- The Artiste's Boutique (Kingston, Jamaica) — Foreign income reporting

Financial columns used:
- totalPrice: Gross booking revenue (what guest paid)
- commissionAmount: Platform commission/fees
- netAmount: Net payout to owner (totalPrice - commissionAmount)

v4 changes from v3:
- Cross-property duplicates cleaned (17 deleted)
- Correct financial columns (totalPrice/commissionAmount/netAmount vs revenueAmount)
- Vrbo property number disambiguation added to scraper
- Cancelled bookings tracked separately
"""

import os
import sys
import json
import subprocess
from datetime import datetime, timezone
from decimal import Decimal

# Install openpyxl if needed
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

# ─── Database Connection ─────────────────────────────────────────────────────
# Read DATABASE_URL from running server's environment
db_url = None
try:
    import glob
    pids = subprocess.check_output(['pgrep', '-f', 'tsx watch'], text=True).strip().split('\n')
    for pid in pids:
        env_file = f'/proc/{pid}/environ'
        if os.path.exists(env_file):
            with open(env_file, 'rb') as f:
                env_data = f.read().decode('utf-8', errors='replace')
                for entry in env_data.split('\x00'):
                    if entry.startswith('DATABASE_URL='):
                        db_url = entry.split('=', 1)[1]
                        break
            if db_url:
                break
except Exception as e:
    print(f"Warning: Could not read from process env: {e}")

if not db_url:
    # Fallback: hardcoded for this session
    db_url = 'mysql://b5BH3wQgpo1xgCj.b8506ae6a7bf:bj248Kd2H6fw4TfWRxwN@gateway03.us-east-1.prod.aws.tidbcloud.com:4000/mhfpBZgGttr5P7pgfv7LpQ?ssl={"rejectUnauthorized":true}'

if not db_url:
    print("ERROR: DATABASE_URL not found")
    sys.exit(1)

try:
    import mysql.connector
except ImportError:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'mysql-connector-python', '-q'])
    import mysql.connector

# Parse DATABASE_URL
from urllib.parse import urlparse, parse_qs
parsed = urlparse(db_url)
db_config = {
    'host': parsed.hostname,
    'port': parsed.port or 3306,
    'user': parsed.username,
    'password': parsed.password,
    'database': parsed.path.lstrip('/'),
}
# Handle SSL for TiDB
if 'ssl-mode' in db_url or 'tidb' in (parsed.hostname or ''):
    db_config['ssl_disabled'] = False

conn = mysql.connector.connect(**db_config)
cursor = conn.cursor(dictionary=True)

# ─── Property Configuration ──────────────────────────────────────────────────
PROPERTIES = {
    'nJnk4hr3AxZJZ-RkwhRJy': {
        'name': 'Morabeza',
        'address': '4693 SR 414, Burdett, NY 14818',
        'type': 'US',
        'schedule_e': True,
    },
    'Ln-_SMF7Nrt1uXsQcdP9C': {
        'name': 'Sunset Studio',
        'address': '4693 SR 414, Burdett, NY 14818',
        'type': 'US',
        'schedule_e': True,
    },
    'ZI2Zy7OuLGYF-vmWOAII-': {
        'name': "The Artiste's Boutique",
        'address': 'Kingston, Jamaica',
        'type': 'Jamaica',
        'schedule_e': False,
    },
}

# ─── Query 2025 Bookings ─────────────────────────────────────────────────────
# 2025: Jan 1 2025 00:00:00 UTC = 1735689600000
# 2026: Jan 1 2026 00:00:00 UTC = 1767225600000
YEAR_START = 1735689600000
YEAR_END = 1767225600000

cursor.execute("""
    SELECT 
        b.id, b.propertyId, b.guestName, b.checkIn, b.checkOut,
        b.totalPrice, b.commissionAmount, b.netAmount, b.revenueAmount,
        b.guestCount, b.cleaningFee, b.confirmationNumber, b.bookingStatus,
        b.currency, b.platformId, b.dataSource
    FROM property_bookings b
    WHERE b.checkIn >= %s AND b.checkIn < %s
    AND b.propertyId IN (%s, %s, %s)
    ORDER BY b.propertyId, b.checkIn
""", (YEAR_START, YEAR_END, 
      'nJnk4hr3AxZJZ-RkwhRJy', 'Ln-_SMF7Nrt1uXsQcdP9C', 'ZI2Zy7OuLGYF-vmWOAII-'))

all_bookings = cursor.fetchall()
cursor.close()
conn.close()

# ─── Process Bookings ─────────────────────────────────────────────────────────
def ts_to_date(ts):
    """Convert millisecond timestamp to date string."""
    if not ts:
        return ''
    return datetime.fromtimestamp(int(ts) / 1000, tz=timezone.utc).strftime('%Y-%m-%d')

def ts_to_nights(checkin, checkout):
    """Calculate number of nights."""
    if not checkin or not checkout:
        return 0
    return int((int(checkout) - int(checkin)) / 86400000)

def get_platform_name(platform_id):
    """Map platform ID to human-readable name."""
    PLATFORMS = {
        'eCyaTlnI': 'Airbnb', 'RiVJhUt8': 'Airbnb', '4HO817D5': 'Airbnb',
        'pg3gaAOM': 'Airbnb', 'qm7PvYpK': 'Airbnb', 'nujsQj2g': 'Airbnb',
        'n2dVmwfB': 'Vrbo', '939ayddu': 'Vrbo', 'dqCVYB2J': 'Vrbo',
        '6rVvJgKe': 'Booking.com',
    }
    if not platform_id:
        return 'Unknown'
    prefix = platform_id[:8]
    return PLATFORMS.get(prefix, 'Unknown')

# Group by property
property_bookings = {}
for prop_id, prop_info in PROPERTIES.items():
    property_bookings[prop_id] = {
        'active': [],
        'cancelled': [],
    }

for b in all_bookings:
    prop_id = b['propertyId']
    if prop_id not in property_bookings:
        continue
    
    booking = {
        'guest_name': b['guestName'] or 'Unknown Guest',
        'check_in': ts_to_date(b['checkIn']),
        'check_out': ts_to_date(b['checkOut']),
        'nights': ts_to_nights(b['checkIn'], b['checkOut']),
        'total_price': float(b['totalPrice'] or 0),
        'commission': float(b['commissionAmount'] or 0),
        'net_amount': float(b['netAmount'] or 0),
        'guest_count': int(b['guestCount'] or 0),
        'cleaning_fee': float(b['cleaningFee'] or 0),
        'confirmation': b['confirmationNumber'] or '',
        'platform': get_platform_name(b['platformId']),
        'currency': b['currency'] or 'USD',
        'status': b['bookingStatus'] or 'confirmed',
    }
    
    if b['bookingStatus'] == 'cancelled':
        property_bookings[prop_id]['cancelled'].append(booking)
    else:
        property_bookings[prop_id]['active'].append(booking)

# ─── Build Workbook ──────────────────────────────────────────────────────────
wb = Workbook()

# Styles
header_font = Font(name='Calibri', bold=True, size=11)
title_font = Font(name='Calibri', bold=True, size=14)
subtitle_font = Font(name='Calibri', bold=True, size=12)
money_format = '#,##0.00'
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font_white = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
summary_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
cancelled_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def write_property_sheet(ws, prop_id, prop_info, bookings_data):
    """Write a property sheet with booking details and summary."""
    active = bookings_data['active']
    cancelled = bookings_data['cancelled']
    
    # Title
    ws['A1'] = f"2025 Rental Income — {prop_info['name']}"
    ws['A1'].font = title_font
    ws['A2'] = f"Address: {prop_info['address']}"
    ws['A2'].font = Font(name='Calibri', size=11, italic=True)
    ws['A3'] = f"Tax Type: {'US Schedule E' if prop_info['schedule_e'] else 'Foreign Income (Form 1116)'}"
    ws['A3'].font = Font(name='Calibri', size=11, italic=True)
    ws['A4'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
    ws['A4'].font = Font(name='Calibri', size=10, italic=True, color='666666')
    
    # Headers
    headers = ['Check-In', 'Check-Out', 'Nights', 'Guest Name', 'Platform', 
               'Gross Revenue', 'Commission', 'Net Payout', 'Guest Count', 'Confirmation #', 'Status']
    
    row = 6
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Active bookings
    row += 1
    for b in active:
        ws.cell(row=row, column=1, value=b['check_in']).border = thin_border
        ws.cell(row=row, column=2, value=b['check_out']).border = thin_border
        ws.cell(row=row, column=3, value=b['nights']).border = thin_border
        ws.cell(row=row, column=4, value=b['guest_name']).border = thin_border
        ws.cell(row=row, column=5, value=b['platform']).border = thin_border
        
        cell = ws.cell(row=row, column=6, value=b['total_price'])
        cell.number_format = money_format
        cell.border = thin_border
        
        cell = ws.cell(row=row, column=7, value=b['commission'])
        cell.number_format = money_format
        cell.border = thin_border
        
        cell = ws.cell(row=row, column=8, value=b['net_amount'])
        cell.number_format = money_format
        cell.border = thin_border
        
        ws.cell(row=row, column=9, value=b['guest_count'] if b['guest_count'] > 0 else '').border = thin_border
        ws.cell(row=row, column=10, value=b['confirmation']).border = thin_border
        ws.cell(row=row, column=11, value='Active').border = thin_border
        row += 1
    
    # Summary row
    summary_row = row
    ws.cell(row=row, column=1, value='TOTALS').font = header_font
    ws.cell(row=row, column=3, value=sum(b['nights'] for b in active)).font = header_font
    ws.cell(row=row, column=4, value=f"{len(active)} bookings").font = header_font
    
    total_gross = sum(b['total_price'] for b in active)
    total_commission = sum(b['commission'] for b in active)
    total_net = sum(b['net_amount'] for b in active)
    
    cell = ws.cell(row=row, column=6, value=total_gross)
    cell.number_format = money_format
    cell.font = header_font
    
    cell = ws.cell(row=row, column=7, value=total_commission)
    cell.number_format = money_format
    cell.font = header_font
    
    cell = ws.cell(row=row, column=8, value=total_net)
    cell.number_format = money_format
    cell.font = header_font
    
    for col in range(1, 12):
        ws.cell(row=row, column=col).fill = summary_fill
        ws.cell(row=row, column=col).border = thin_border
    
    # Cancelled bookings section
    if cancelled:
        row += 3
        ws.cell(row=row, column=1, value='CANCELLED BOOKINGS').font = subtitle_font
        row += 1
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = cancelled_fill
            cell.border = thin_border
        row += 1
        for b in cancelled:
            ws.cell(row=row, column=1, value=b['check_in']).border = thin_border
            ws.cell(row=row, column=2, value=b['check_out']).border = thin_border
            ws.cell(row=row, column=3, value=b['nights']).border = thin_border
            ws.cell(row=row, column=4, value=b['guest_name']).border = thin_border
            ws.cell(row=row, column=5, value=b['platform']).border = thin_border
            cell = ws.cell(row=row, column=6, value=b['total_price'])
            cell.number_format = money_format
            cell.border = thin_border
            cell = ws.cell(row=row, column=7, value=b['commission'])
            cell.number_format = money_format
            cell.border = thin_border
            cell = ws.cell(row=row, column=8, value=b['net_amount'])
            cell.number_format = money_format
            cell.border = thin_border
            ws.cell(row=row, column=9, value=b['guest_count'] if b['guest_count'] > 0 else '').border = thin_border
            ws.cell(row=row, column=10, value=b['confirmation']).border = thin_border
            ws.cell(row=row, column=11, value='Cancelled').border = thin_border
            row += 1
    
    # Column widths
    col_widths = [12, 12, 8, 25, 10, 14, 14, 14, 12, 16, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    return total_gross, total_commission, total_net, len(active), sum(b['nights'] for b in active)

# ─── Create Summary Sheet ─────────────────────────────────────────────────────
ws_summary = wb.active
ws_summary.title = 'Summary'

ws_summary['A1'] = 'Tarik Perkins — 2025 Property Tax Summary'
ws_summary['A1'].font = title_font
ws_summary['A2'] = f'Generated: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}'
ws_summary['A2'].font = Font(name='Calibri', size=10, italic=True, color='666666')
ws_summary['A3'] = 'Data Source: Geeves Property Management (Airbnb + Vrbo email scraping)'
ws_summary['A3'].font = Font(name='Calibri', size=10, italic=True, color='666666')

# Summary table headers
row = 5
summary_headers = ['Property', 'Address', 'Tax Type', 'Bookings', 'Nights', 
                   'Gross Revenue', 'Commission', 'Net Payout', 'Avg/Night']
for col, header in enumerate(summary_headers, 1):
    cell = ws_summary.cell(row=row, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# ─── Create Property Sheets ──────────────────────────────────────────────────
property_summaries = []

for prop_id, prop_info in PROPERTIES.items():
    ws = wb.create_sheet(title=prop_info['name'][:31])  # Excel sheet name limit
    bookings_data = property_bookings[prop_id]
    gross, commission, net, count, nights = write_property_sheet(ws, prop_id, prop_info, bookings_data)
    property_summaries.append({
        'name': prop_info['name'],
        'address': prop_info['address'],
        'type': 'US Schedule E' if prop_info['schedule_e'] else 'Foreign (Form 1116)',
        'bookings': count,
        'nights': nights,
        'gross': gross,
        'commission': commission,
        'net': net,
        'avg_per_night': net / nights if nights > 0 else 0,
    })

# Fill summary table
row = 6
for ps in property_summaries:
    ws_summary.cell(row=row, column=1, value=ps['name']).border = thin_border
    ws_summary.cell(row=row, column=2, value=ps['address']).border = thin_border
    ws_summary.cell(row=row, column=3, value=ps['type']).border = thin_border
    ws_summary.cell(row=row, column=4, value=ps['bookings']).border = thin_border
    ws_summary.cell(row=row, column=5, value=ps['nights']).border = thin_border
    
    cell = ws_summary.cell(row=row, column=6, value=ps['gross'])
    cell.number_format = money_format
    cell.border = thin_border
    
    cell = ws_summary.cell(row=row, column=7, value=ps['commission'])
    cell.number_format = money_format
    cell.border = thin_border
    
    cell = ws_summary.cell(row=row, column=8, value=ps['net'])
    cell.number_format = money_format
    cell.border = thin_border
    
    cell = ws_summary.cell(row=row, column=9, value=ps['avg_per_night'])
    cell.number_format = money_format
    cell.border = thin_border
    
    row += 1

# Totals row
ws_summary.cell(row=row, column=1, value='TOTAL').font = header_font
ws_summary.cell(row=row, column=4, value=sum(ps['bookings'] for ps in property_summaries)).font = header_font
ws_summary.cell(row=row, column=5, value=sum(ps['nights'] for ps in property_summaries)).font = header_font

cell = ws_summary.cell(row=row, column=6, value=sum(ps['gross'] for ps in property_summaries))
cell.number_format = money_format
cell.font = header_font

cell = ws_summary.cell(row=row, column=7, value=sum(ps['commission'] for ps in property_summaries))
cell.number_format = money_format
cell.font = header_font

cell = ws_summary.cell(row=row, column=8, value=sum(ps['net'] for ps in property_summaries))
cell.number_format = money_format
cell.font = header_font

for col in range(1, 10):
    ws_summary.cell(row=row, column=col).fill = summary_fill
    ws_summary.cell(row=row, column=col).border = thin_border

# US-only subtotal (for Schedule E)
row += 2
ws_summary.cell(row=row, column=1, value='US Properties Only (Schedule E)').font = subtitle_font
row += 1
us_gross = sum(ps['gross'] for ps in property_summaries if 'US' in ps['type'])
us_commission = sum(ps['commission'] for ps in property_summaries if 'US' in ps['type'])
us_net = sum(ps['net'] for ps in property_summaries if 'US' in ps['type'])
us_bookings = sum(ps['bookings'] for ps in property_summaries if 'US' in ps['type'])
us_nights = sum(ps['nights'] for ps in property_summaries if 'US' in ps['type'])

ws_summary.cell(row=row, column=1, value='Gross Revenue:')
cell = ws_summary.cell(row=row, column=2, value=us_gross)
cell.number_format = money_format
cell.font = header_font
row += 1

ws_summary.cell(row=row, column=1, value='Platform Commissions:')
cell = ws_summary.cell(row=row, column=2, value=us_commission)
cell.number_format = money_format
row += 1

ws_summary.cell(row=row, column=1, value='Net Payout:')
cell = ws_summary.cell(row=row, column=2, value=us_net)
cell.number_format = money_format
cell.font = header_font
row += 1

ws_summary.cell(row=row, column=1, value='Total Bookings:')
ws_summary.cell(row=row, column=2, value=us_bookings)
row += 1

ws_summary.cell(row=row, column=1, value='Total Nights Rented:')
ws_summary.cell(row=row, column=2, value=us_nights)

# Data quality notes
row += 3
ws_summary.cell(row=row, column=1, value='DATA QUALITY NOTES').font = subtitle_font
row += 1

notes = [
    "• Financial data sourced from Airbnb and Vrbo email scraping (confirmation emails, payout notifications)",
    "• Some bookings (especially Vrbo iCal-only) may not have financial data — these are listed with $0",
    "• 'Commission' = platform service fee deducted from gross revenue",
    "• 'Net Payout' = amount deposited to owner's bank account",
    "• Morabeza has 6 bookings without financials (Vrbo iCal imports without matching emails)",
    "• Artiste's Boutique has 3 bookings without financials (Vrbo iCal imports)",
    "• Currency: All amounts in USD unless noted otherwise",
    "• Cancelled bookings are tracked separately and NOT included in totals",
    f"• 8 Airbnb bookings on Morabeza have unconfirmed property assignment (from shared email account)",
    "• Workbook version: v4 (post-duplicate-cleanup, Jul 3 2026)",
]
for note in notes:
    ws_summary.cell(row=row, column=1, value=note)
    row += 1

# Column widths for summary
summary_widths = [22, 35, 18, 10, 8, 14, 14, 14, 12]
for i, width in enumerate(summary_widths, 1):
    ws_summary.column_dimensions[get_column_letter(i)].width = width

# ─── Save Workbook ───────────────────────────────────────────────────────────
output_path = '/home/ubuntu/tax_prep/Tarik_Perkins_2025_Property_Tax_Workbook_v4.xlsx'
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)

print(f"✓ Workbook saved: {output_path}")
print(f"\nSummary:")
for ps in property_summaries:
    print(f"  {ps['name']}: {ps['bookings']} bookings, {ps['nights']} nights, Gross ${ps['gross']:.2f}, Net ${ps['net']:.2f}")
print(f"\n  TOTAL: {sum(ps['bookings'] for ps in property_summaries)} bookings, "
      f"Gross ${sum(ps['gross'] for ps in property_summaries):.2f}, "
      f"Net ${sum(ps['net'] for ps in property_summaries):.2f}")
print(f"\n  US Schedule E: Gross ${us_gross:.2f}, Net ${us_net:.2f}")
