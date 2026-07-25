import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# Load all 3 PayPal CSVs
pp_2024 = pd.read_csv('/home/ubuntu/tax_prep/gdrive_data/paypal_new/PayPal 2024.CSV', encoding='utf-8-sig')
pp_2025 = pd.read_csv('/home/ubuntu/tax_prep/gdrive_data/paypal_new/PayPal 2025.CSV', encoding='utf-8-sig')
pp_2026 = pd.read_csv('/home/ubuntu/tax_prep/gdrive_data/paypal_new/PayPal Jan - June 2026.CSV', encoding='utf-8-sig')

# Combine all
pp_all = pd.concat([pp_2024, pp_2025, pp_2026], ignore_index=True)

# Parse amounts
def parse_amount(val):
    if pd.isna(val) or val == '':
        return 0.0
    return float(str(val).replace(',', ''))

pp_all['Gross_num'] = pp_all['Gross'].apply(parse_amount)
pp_all['Fee_num'] = pp_all['Fee'].apply(parse_amount)
pp_all['Net_num'] = pp_all['Net'].apply(parse_amount)

# Parse dates
pp_all['Date_parsed'] = pd.to_datetime(pp_all['Date'], format='%m/%d/%Y', errors='coerce')

# Auto-classify transactions
def classify_transaction(row):
    """Auto-classify PayPal transactions for QBO export"""
    item = str(row.get('Item Title', '')).lower() if pd.notna(row.get('Item Title')) else ''
    name = str(row.get('Name', '')).lower() if pd.notna(row.get('Name')) else ''
    from_email = str(row.get('From Email Address', '')).lower() if pd.notna(row.get('From Email Address')) else ''
    to_email = str(row.get('To Email Address', '')).lower() if pd.notna(row.get('To Email Address')) else ''
    txn_type = str(row.get('Type', '')) if pd.notna(row.get('Type')) else ''
    status = str(row.get('Status', '')) if pd.notna(row.get('Status')) else ''
    gross = row.get('Gross_num', 0)
    balance_impact = str(row.get('Balance Impact', '')) if pd.notna(row.get('Balance Impact')) else ''
    
    # STR Income - Booking.com guest payments (from guest.booking.com emails)
    if 'booking.com' in from_email or 'guest.booking' in from_email:
        if txn_type == 'Invoice Sent':
            if status == 'Paid':
                return 'STR Income - Booking.com Invoice (Paid)', 'Artiste\'s Boutique'
            elif status == 'Canceled':
                return 'STR Income - Booking.com Invoice (Cancelled)', 'Artiste\'s Boutique'
            else:
                return 'STR Income - Booking.com Invoice (Unpaid)', 'Artiste\'s Boutique'
        return 'STR Income - Booking.com', 'Artiste\'s Boutique'
    
    # STR Income - Express Checkout from booking.com guests
    if txn_type == 'Express Checkout Payment' and gross > 100:
        if 'booking.com' in from_email or 'guest.booking' in from_email:
            return 'STR Income - Booking.com (PayPal)', 'Artiste\'s Boutique'
        # Check item for STR keywords
        if 'night' in item or 'stay' in item or 'artiste' in item or 'boutique' in item or 'lodging' in item or 'cleaning' in item:
            return 'STR Income - Guest Payment', 'Artiste\'s Boutique'
        if 'invoice' in item:
            return 'STR Income - Guest Payment', 'Artiste\'s Boutique'
        return 'STR Income - Guest Payment', 'Artiste\'s Boutique'
    
    # Invoice sent to booking.com guests
    if txn_type == 'Invoice Sent':
        if 'booking' in from_email or 'nightly' in item or 'cleaning' in item or 'lodging' in item:
            if status == 'Paid':
                return 'STR Income - Booking.com Invoice (Paid)', 'Artiste\'s Boutique'
            elif 'cancel' in status.lower():
                return 'STR Income - Booking.com Invoice (Cancelled)', 'Artiste\'s Boutique'
            else:
                return 'STR Income - Booking.com Invoice (Unpaid)', 'Artiste\'s Boutique'
        # Car rental invoices
        if 'rental' in item or 'chevy' in item or 'trax' in item or 'speeding' in item:
            return 'Other Income - Car Rental', 'Vehicle'
        return 'STR Income - Invoice', 'Unclassified'
    
    # Subscriptions
    if 'disney' in item:
        return 'Personal - Subscription', 'Disney+'
    if 'grindr' in item:
        return 'Personal - Subscription', 'Grindr'
    if 'impulse' in item:
        return 'Personal - Subscription', 'Impulse Brain Training'
    if 'stir' in item:
        return 'Personal - Subscription', 'Stir Dating'
    if 'max:' in item or 'hbo' in item:
        return 'Personal - Subscription', 'Max/HBO'
    if 'youtube' in item:
        return 'Personal - Subscription', 'YouTube'
    if 'chatgpt' in item or 'openai' in item:
        return 'Business - Software', 'ChatGPT/OpenAI'
    if 'canva' in item:
        return 'Business - Software', 'Canva'
    if 'notion' in item:
        return 'Business - Software', 'Notion'
    if 'google' in item and 'storage' in item:
        return 'Business - Software', 'Google Storage'
    if 'adobe' in item:
        return 'Business - Software', 'Adobe'
    if 'manus' in item:
        return 'Business - Software', 'Manus AI'
    
    # Transfers
    if 'bank' in txn_type.lower() and 'transfer' in txn_type.lower():
        return 'Transfer', 'Bank Transfer'
    if txn_type == 'General Withdrawal':
        return 'Transfer', 'Withdrawal to Bank'
    if txn_type == 'General Authorization':
        return 'Transfer', 'Authorization Hold'
    
    # Refunds
    if 'reversal' in txn_type.lower() or 'refund' in txn_type.lower():
        return 'Refund', ''
    
    # Purchases via PayPal balance
    if txn_type == 'Bank Deposit to PP Account':
        return 'Personal - Purchase/Subscription', ''
    if txn_type == 'General Card Deposit':
        return 'Personal - Purchase', ''
    
    # Negative Express Checkout = outgoing payment
    if txn_type == 'Express Checkout Payment' and gross < 0:
        return 'Personal - Purchase', ''
    
    # Catch-all for debits
    if balance_impact == 'Debit' and gross < 0:
        return 'Personal - Purchase', ''
    
    return 'Unclassified', ''

# Apply classification
pp_all[['Category', 'Sub_Category']] = pp_all.apply(
    lambda row: pd.Series(classify_transaction(row)), axis=1
)

# Print summary by year and category
for year in [2024, 2025, 2026]:
    year_data = pp_all[pp_all['Date_parsed'].dt.year == year]
    if len(year_data) == 0:
        continue
    print(f"\n{'='*60}")
    print(f"{year} PAYPAL SUMMARY")
    print(f"{'='*60}")
    
    # Credits
    credits = year_data[year_data['Balance Impact'] == 'Credit']
    if len(credits) > 0:
        print(f"\n  INCOME (Credits): {len(credits)} txns, ${credits['Gross_num'].sum():,.2f}")
        for cat, grp in credits.groupby('Category'):
            print(f"    {cat}: {len(grp)} txns, ${grp['Gross_num'].sum():,.2f}")
    
    # Debits
    debits = year_data[year_data['Balance Impact'] == 'Debit']
    if len(debits) > 0:
        print(f"\n  EXPENSES (Debits): {len(debits)} txns, ${debits['Gross_num'].sum():,.2f}")
        for cat, grp in debits.groupby('Category'):
            print(f"    {cat}: {len(grp)} txns, ${grp['Gross_num'].sum():,.2f}")
    
    # Memo/no impact
    memo = year_data[~year_data['Balance Impact'].isin(['Credit', 'Debit'])]
    if len(memo) > 0:
        print(f"\n  NO BALANCE IMPACT (Memo): {len(memo)} txns")
        for cat, grp in memo.groupby('Category'):
            print(f"    {cat}: {len(grp)} txns")

# Now create the workbook tab
print(f"\n\n{'='*60}")
print("CREATING PAYPAL RECONCILIATION TAB IN WORKBOOK...")
print(f"{'='*60}")

# Load existing workbook
wb = openpyxl.load_workbook('/home/ubuntu/tax_prep/Tarik_Perkins_2025_Property_Tax_Workbook_v5.xlsx')

# Remove existing PayPal tab if present
if 'PayPal' in wb.sheetnames:
    del wb['PayPal']

ws = wb.create_sheet('PayPal')

# Styles
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font_white = Font(bold=True, size=11, color='FFFFFF')
income_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
expense_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
str_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Title
ws['A1'] = 'PayPal Reconciliation - All Transactions (2024-2026)'
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:L1')

# Summary section
ws['A3'] = '2024 SUMMARY'
ws['A3'].font = Font(bold=True, size=12)

# 2024 summary
year_2024 = pp_all[pp_all['Date_parsed'].dt.year == 2024]
credits_2024 = year_2024[year_2024['Balance Impact'] == 'Credit']
debits_2024 = year_2024[year_2024['Balance Impact'] == 'Debit']

ws['A4'] = 'Total Income (Credits):'
ws['B4'] = credits_2024['Gross_num'].sum()
ws['B4'].number_format = '$#,##0.00'
ws['A5'] = 'Total Expenses (Debits):'
ws['B5'] = debits_2024['Gross_num'].sum()
ws['B5'].number_format = '$#,##0.00'
ws['A6'] = 'STR Income (Booking.com):'
str_income = credits_2024[credits_2024['Category'].str.contains('STR', na=False)]
ws['B6'] = str_income['Gross_num'].sum()
ws['B6'].number_format = '$#,##0.00'
ws['B6'].fill = str_fill
ws['A7'] = 'QBO PayPal STR Amount:'
ws['B7'] = 1638.00
ws['B7'].number_format = '$#,##0.00'
ws['A8'] = 'Difference:'
ws['B8'] = str_income['Gross_num'].sum() - 1638.00
ws['B8'].number_format = '$#,##0.00'

# Headers for transaction detail
headers = ['Date', 'Name', 'Type', 'Status', 'Gross', 'Fee', 'Net', 
           'From Email', 'Item Title', 'Category', 'Sub-Category', 
           'Balance Impact', 'Transaction ID', 'Year']
row_start = 11

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=row_start, column=col, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Sort by date
pp_sorted = pp_all.sort_values('Date_parsed').copy()

# Write data rows
for i, (_, row) in enumerate(pp_sorted.iterrows()):
    r = row_start + 1 + i
    
    ws.cell(row=r, column=1, value=row['Date_parsed'].strftime('%m/%d/%Y') if pd.notna(row['Date_parsed']) else '')
    ws.cell(row=r, column=2, value=str(row['Name']) if pd.notna(row['Name']) else '')
    ws.cell(row=r, column=3, value=str(row['Type']) if pd.notna(row['Type']) else '')
    ws.cell(row=r, column=4, value=str(row['Status']) if pd.notna(row['Status']) else '')
    
    gross_cell = ws.cell(row=r, column=5, value=row['Gross_num'])
    gross_cell.number_format = '$#,##0.00'
    
    fee_cell = ws.cell(row=r, column=6, value=row['Fee_num'])
    fee_cell.number_format = '$#,##0.00'
    
    net_cell = ws.cell(row=r, column=7, value=row['Net_num'])
    net_cell.number_format = '$#,##0.00'
    
    ws.cell(row=r, column=8, value=str(row['From Email Address']) if pd.notna(row['From Email Address']) else '')
    ws.cell(row=r, column=9, value=str(row['Item Title']) if pd.notna(row['Item Title']) else '')
    ws.cell(row=r, column=10, value=row['Category'])
    ws.cell(row=r, column=11, value=row['Sub_Category'])
    ws.cell(row=r, column=12, value=str(row['Balance Impact']) if pd.notna(row['Balance Impact']) else '')
    ws.cell(row=r, column=13, value=str(row['Transaction ID']) if pd.notna(row['Transaction ID']) else '')
    ws.cell(row=r, column=14, value=row['Date_parsed'].year if pd.notna(row['Date_parsed']) else '')
    
    # Color code rows
    balance_impact = str(row['Balance Impact']) if pd.notna(row['Balance Impact']) else ''
    category = row['Category']
    
    fill = None
    if 'STR' in category:
        fill = str_fill
    elif balance_impact == 'Credit':
        fill = income_fill
    elif balance_impact == 'Debit':
        fill = expense_fill
    
    if fill:
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).fill = fill
    
    # Add border
    for col in range(1, len(headers) + 1):
        ws.cell(row=r, column=col).border = thin_border

# Set column widths
widths = [12, 30, 22, 12, 12, 10, 12, 35, 40, 28, 20, 14, 20, 6]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Add filter
ws.auto_filter.ref = f"A{row_start}:{get_column_letter(len(headers))}{row_start + len(pp_sorted)}"

# Save
output_path = '/home/ubuntu/tax_prep/Tarik_Perkins_2025_Property_Tax_Workbook_v6.xlsx'
wb.save(output_path)
print(f"\nWorkbook saved to: {output_path}")
print(f"PayPal tab has {len(pp_sorted)} transactions with auto-classification")
print(f"\nColor coding:")
print(f"  Blue = STR Income (Booking.com)")
print(f"  Green = Other Income (Credits)")
print(f"  Red = Expenses (Debits)")
print(f"  White = Memo/No balance impact")
print(f"\nYou can override the 'Category' and 'Sub-Category' columns to reclassify any transaction")
