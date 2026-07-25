"""
Build the 2025 US Property Tax Preparation Workbook.
Sheets:
1. US Income Dashboard - Monthly pivot by property x platform
2. US Expenses - Categorized property-related expenses
3. US P&L Summary - Income minus expenses per property (Schedule E ready)
4. Jamaica Income - Jamaica property income (Artiste's Boutique)
5. All Airbnb Payouts Detail - Raw payout data for reference
"""
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# Load parsed data
with open('parsed/airbnb_payouts_parsed.json') as f:
    airbnb_payouts = json.load(f)

with open('parsed/property_expenses_parsed.json') as f:
    property_expenses = json.load(f)

with open('raw_data/db_bookings_2025_full.json') as f:
    db_bookings = json.load(f)['bookings']

# ============================================================
# NORMALIZE PROPERTY NAMES
# ============================================================
def normalize_property(name):
    if not name:
        return 'Unknown'
    name_lower = name.lower()
    if 'morabeza' in name_lower or 'seneca haven' in name_lower:
        return 'Morabeza (Studio)'
    elif 'sunset' in name_lower or 'seneca sunsets' in name_lower:
        return 'Sunset Studio'
    elif 'artis' in name_lower or 'bohemian' in name_lower or 'boutique' in name_lower:
        return "The Artiste's Boutique"
    elif 'penthouse' in name_lower or 'two-bedroom' in name_lower:
        return 'Two-Bedroom Penthouse (Retired)'
    return name

def get_month(date_str):
    """Extract month number from YYYY-MM-DD date string."""
    if not date_str or len(date_str) < 7:
        return None
    try:
        return int(date_str[5:7])
    except:
        return None

def is_jamaica_property(prop_name):
    """Determine if a property is in Jamaica."""
    return prop_name == "The Artiste's Boutique"

# ============================================================
# PROCESS AIRBNB INCOME DATA
# ============================================================
income_records = []

for payout in airbnb_payouts:
    if 'error' in payout:
        continue
    for item in payout.get('line_items', []):
        prop = normalize_property(item.get('property_name', ''))
        # Skip "Pass Through Tot" as it's tax pass-through, not rental income
        desc = item.get('description', '')
        is_tax_passthrough = 'pass through' in desc.lower() or 'tot' in desc.lower()
        
        record = {
            'date': item.get('check_in') or payout.get('payout_date', ''),
            'payout_date': payout.get('payout_date', ''),
            'property': prop,
            'platform': 'Airbnb',
            'guest': item.get('guest_name', ''),
            'check_in': item.get('check_in', ''),
            'check_out': item.get('check_out', ''),
            'amount': item.get('amount', 0),
            'type': 'tax_passthrough' if is_tax_passthrough else 'rental_income',
            'payout_id': payout.get('payout_id', ''),
            'confirmation_code': item.get('confirmation_code', ''),
            'listing_id': item.get('listing_id', ''),
            'month': get_month(item.get('check_in') or payout.get('payout_date', ''))
        }
        income_records.append(record)

# Also add DB bookings that have financial data (non-Airbnb or enriched)
for booking in db_bookings:
    if booking.get('revenue') and float(booking['revenue']) > 0:
        # Check if this is already captured in Airbnb payouts
        # Skip if code matches an existing Airbnb record
        code = booking.get('code', '')
        already_exists = any(r['confirmation_code'] == code for r in income_records if code and r['confirmation_code'])
        if already_exists:
            continue
        
        platform = 'Unknown'
        cal_name = (booking.get('calendarName') or '').lower()
        if 'airbnb' in cal_name:
            platform = 'Airbnb'
        elif 'vrbo' in cal_name:
            platform = 'Vrbo'
        elif 'booking' in cal_name:
            platform = 'Booking.com'
        
        record = {
            'date': booking.get('checkIn', ''),
            'payout_date': '',
            'property': normalize_property(booking.get('property', '')),
            'platform': platform,
            'guest': booking.get('guest', ''),
            'check_in': booking.get('checkIn', ''),
            'check_out': booking.get('checkOut', ''),
            'amount': float(booking.get('net') or booking.get('revenue') or 0),
            'type': 'rental_income',
            'payout_id': '',
            'confirmation_code': code,
            'listing_id': '',
            'month': get_month(booking.get('checkIn', ''))
        }
        income_records.append(record)

# Separate US and Jamaica income
us_income = [r for r in income_records if not is_jamaica_property(r['property'])]
jamaica_income = [r for r in income_records if is_jamaica_property(r['property'])]

print(f"Total income records: {len(income_records)}")
print(f"  US: {len(us_income)}")
print(f"  Jamaica: {len(jamaica_income)}")

# ============================================================
# BUILD WORKBOOK
# ============================================================
wb = Workbook()

# Styles
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
subheader_font = Font(bold=True, size=11)
subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
money_format = '#,##0.00'
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

def auto_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

# ============================================================
# SHEET 1: US Income Dashboard
# ============================================================
ws1 = wb.active
ws1.title = "US Income Dashboard"

# Title
ws1['A1'] = "2025 US Property Income Dashboard"
ws1['A1'].font = Font(bold=True, size=16)
ws1['A2'] = f"Prepared: {datetime.now().strftime('%B %d, %Y')}"
ws1['A3'] = "Owner: Tarik Perkins"
ws1['A4'] = "Properties: Upstate New York"

# Monthly pivot table
months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
us_properties = sorted(set(r['property'] for r in us_income if r['property'] != 'Unknown'))
if not us_properties:
    us_properties = ['Morabeza (Studio)', 'Sunset Studio']

row = 6
ws1.cell(row=row, column=1, value="Property / Month")
for i, m in enumerate(months):
    ws1.cell(row=row, column=i+2, value=m)
ws1.cell(row=row, column=14, value="TOTAL")
style_header_row(ws1, row, 14)

for prop in us_properties:
    row += 1
    ws1.cell(row=row, column=1, value=prop)
    ws1.cell(row=row, column=1).font = Font(bold=True)
    prop_total = 0
    for m_idx in range(1, 13):
        month_income = sum(r['amount'] for r in us_income 
                         if r['property'] == prop and r['month'] == m_idx and r['type'] == 'rental_income')
        ws1.cell(row=row, column=m_idx+1, value=month_income if month_income > 0 else None)
        ws1.cell(row=row, column=m_idx+1).number_format = money_format
        prop_total += month_income
    ws1.cell(row=row, column=14, value=prop_total)
    ws1.cell(row=row, column=14).number_format = money_format
    ws1.cell(row=row, column=14).font = Font(bold=True)

# Tax pass-through row
row += 1
ws1.cell(row=row, column=1, value="Tax Pass-Through (TOT)")
ws1.cell(row=row, column=1).font = Font(italic=True)
tot_total = 0
for m_idx in range(1, 13):
    month_tot = sum(r['amount'] for r in us_income 
                   if r['month'] == m_idx and r['type'] == 'tax_passthrough' and not is_jamaica_property(r['property']))
    ws1.cell(row=row, column=m_idx+1, value=month_tot if month_tot > 0 else None)
    ws1.cell(row=row, column=m_idx+1).number_format = money_format
    tot_total += month_tot
ws1.cell(row=row, column=14, value=tot_total)
ws1.cell(row=row, column=14).number_format = money_format

# Total row
row += 1
ws1.cell(row=row, column=1, value="TOTAL US INCOME")
ws1.cell(row=row, column=1).font = Font(bold=True, size=12)
grand_total = 0
for m_idx in range(1, 13):
    month_all = sum(r['amount'] for r in us_income if r['month'] == m_idx and r['type'] == 'rental_income')
    ws1.cell(row=row, column=m_idx+1, value=month_all if month_all > 0 else None)
    ws1.cell(row=row, column=m_idx+1).number_format = money_format
    ws1.cell(row=row, column=m_idx+1).font = Font(bold=True)
    grand_total += month_all
ws1.cell(row=row, column=14, value=grand_total)
ws1.cell(row=row, column=14).number_format = money_format
ws1.cell(row=row, column=14).font = Font(bold=True, size=12)

# Platform breakdown section
row += 3
ws1.cell(row=row, column=1, value="Income by Platform")
ws1.cell(row=row, column=1).font = Font(bold=True, size=14)
row += 1
headers = ['Platform', 'Bookings', 'Total Income', 'Avg per Booking']
for i, h in enumerate(headers):
    ws1.cell(row=row, column=i+1, value=h)
style_header_row(ws1, row, 4)

platforms = {}
for r in us_income:
    if r['type'] != 'rental_income':
        continue
    p = r['platform']
    if p not in platforms:
        platforms[p] = {'count': 0, 'total': 0}
    platforms[p]['count'] += 1
    platforms[p]['total'] += r['amount']

for platform, data in sorted(platforms.items()):
    row += 1
    ws1.cell(row=row, column=1, value=platform)
    ws1.cell(row=row, column=2, value=data['count'])
    ws1.cell(row=row, column=3, value=data['total'])
    ws1.cell(row=row, column=3).number_format = money_format
    ws1.cell(row=row, column=4, value=data['total'] / data['count'] if data['count'] > 0 else 0)
    ws1.cell(row=row, column=4).number_format = money_format

# Detailed booking list
row += 3
ws1.cell(row=row, column=1, value="Detailed Booking List")
ws1.cell(row=row, column=1).font = Font(bold=True, size=14)
row += 1
detail_headers = ['Date', 'Property', 'Platform', 'Guest', 'Check-In', 'Check-Out', 'Amount', 'Payout ID', 'Conf. Code']
for i, h in enumerate(detail_headers):
    ws1.cell(row=row, column=i+1, value=h)
style_header_row(ws1, row, len(detail_headers))

for r in sorted(us_income, key=lambda x: x.get('date', '')):
    if r['type'] != 'rental_income':
        continue
    row += 1
    ws1.cell(row=row, column=1, value=r['date'])
    ws1.cell(row=row, column=2, value=r['property'])
    ws1.cell(row=row, column=3, value=r['platform'])
    ws1.cell(row=row, column=4, value=r['guest'])
    ws1.cell(row=row, column=5, value=r['check_in'])
    ws1.cell(row=row, column=6, value=r['check_out'])
    ws1.cell(row=row, column=7, value=r['amount'])
    ws1.cell(row=row, column=7).number_format = money_format
    ws1.cell(row=row, column=8, value=r['payout_id'])
    ws1.cell(row=row, column=9, value=r['confirmation_code'])

auto_width(ws1)

# ============================================================
# SHEET 2: US Expenses
# ============================================================
ws2 = wb.create_sheet("US Expenses")

ws2['A1'] = "2025 US Property Expenses"
ws2['A1'].font = Font(bold=True, size=16)
ws2['A2'] = f"Sources: Amazon, Walmart, Venmo (from tariqp@gmail.com)"
ws2['A3'] = "Note: Only property-related expenses are included. Personal purchases excluded."

row = 5
expense_headers = ['Date', 'Vendor', 'Category', 'Description', 'Amount', 'Property', 'Source']
for i, h in enumerate(expense_headers):
    ws2.cell(row=row, column=i+1, value=h)
style_header_row(ws2, row, len(expense_headers))

# Sort expenses by date
sorted_expenses = sorted(property_expenses, key=lambda x: x.get('date', ''))
total_expenses = 0

for e in sorted_expenses:
    row += 1
    ws2.cell(row=row, column=1, value=e.get('date', ''))
    ws2.cell(row=row, column=2, value=e.get('vendor', ''))
    ws2.cell(row=row, column=3, value=e.get('category', ''))
    ws2.cell(row=row, column=4, value=e.get('items_description', '')[:80])
    amt = e.get('total_amount', 0)
    ws2.cell(row=row, column=5, value=amt)
    ws2.cell(row=row, column=5).number_format = money_format
    ws2.cell(row=row, column=6, value=e.get('likely_property', 'Unknown'))
    ws2.cell(row=row, column=7, value=e.get('_source', ''))
    total_expenses += amt

# Total row
row += 2
ws2.cell(row=row, column=4, value="TOTAL PROPERTY EXPENSES")
ws2.cell(row=row, column=4).font = Font(bold=True)
ws2.cell(row=row, column=5, value=total_expenses)
ws2.cell(row=row, column=5).number_format = money_format
ws2.cell(row=row, column=5).font = Font(bold=True)

# Category summary
row += 3
ws2.cell(row=row, column=1, value="Expense Summary by Category")
ws2.cell(row=row, column=1).font = Font(bold=True, size=14)
row += 1
cat_headers = ['Category', 'Count', 'Total Amount']
for i, h in enumerate(cat_headers):
    ws2.cell(row=row, column=i+1, value=h)
style_header_row(ws2, row, 3)

by_cat = {}
for e in sorted_expenses:
    cat = e.get('category', 'other')
    if cat not in by_cat:
        by_cat[cat] = {'count': 0, 'total': 0}
    by_cat[cat]['count'] += 1
    by_cat[cat]['total'] += e.get('total_amount', 0)

for cat, data in sorted(by_cat.items(), key=lambda x: -x[1]['total']):
    row += 1
    ws2.cell(row=row, column=1, value=cat.title())
    ws2.cell(row=row, column=2, value=data['count'])
    ws2.cell(row=row, column=3, value=data['total'])
    ws2.cell(row=row, column=3).number_format = money_format

auto_width(ws2)

# ============================================================
# SHEET 3: US P&L Summary
# ============================================================
ws3 = wb.create_sheet("US P&L Summary")

ws3['A1'] = "2025 US Property Profit & Loss Summary"
ws3['A1'].font = Font(bold=True, size=16)
ws3['A2'] = "Schedule E Ready — Short-Term Rental Properties"
ws3['A3'] = f"Prepared: {datetime.now().strftime('%B %d, %Y')}"

row = 5
pnl_headers = ['Property', 'Gross Rental Income', 'Tax Pass-Through', 'Total Expenses', 'Net Profit/(Loss)']
for i, h in enumerate(pnl_headers):
    ws3.cell(row=row, column=i+1, value=h)
style_header_row(ws3, row, len(pnl_headers))

# Calculate P&L per property
grand_income = 0
grand_tot = 0
grand_expenses = 0
grand_net = 0

for prop in us_properties:
    row += 1
    prop_income = sum(r['amount'] for r in us_income if r['property'] == prop and r['type'] == 'rental_income')
    prop_tot = sum(r['amount'] for r in us_income if r['property'] == prop and r['type'] == 'tax_passthrough')
    # Allocate expenses: if property-specific, use that; otherwise split "General" evenly
    prop_specific_exp = sum(e.get('total_amount', 0) for e in sorted_expenses 
                          if normalize_property(e.get('likely_property', '')) == prop)
    general_exp = sum(e.get('total_amount', 0) for e in sorted_expenses 
                     if e.get('likely_property', '') in ['General (all properties)', 'General', 'Unknown', ''])
    # Split general expenses evenly among US properties
    prop_general_share = general_exp / len(us_properties) if us_properties else 0
    prop_expenses = prop_specific_exp + prop_general_share
    
    net = prop_income - prop_expenses
    
    ws3.cell(row=row, column=1, value=prop)
    ws3.cell(row=row, column=1).font = Font(bold=True)
    ws3.cell(row=row, column=2, value=prop_income)
    ws3.cell(row=row, column=2).number_format = money_format
    ws3.cell(row=row, column=3, value=prop_tot)
    ws3.cell(row=row, column=3).number_format = money_format
    ws3.cell(row=row, column=4, value=prop_expenses)
    ws3.cell(row=row, column=4).number_format = money_format
    ws3.cell(row=row, column=5, value=net)
    ws3.cell(row=row, column=5).number_format = money_format
    ws3.cell(row=row, column=5).font = Font(bold=True, color="006400" if net >= 0 else "8B0000")
    
    grand_income += prop_income
    grand_tot += prop_tot
    grand_expenses += prop_expenses
    grand_net += net

# Totals
row += 1
ws3.cell(row=row, column=1, value="TOTAL")
ws3.cell(row=row, column=1).font = Font(bold=True, size=12)
ws3.cell(row=row, column=2, value=grand_income)
ws3.cell(row=row, column=2).number_format = money_format
ws3.cell(row=row, column=2).font = Font(bold=True)
ws3.cell(row=row, column=3, value=grand_tot)
ws3.cell(row=row, column=3).number_format = money_format
ws3.cell(row=row, column=3).font = Font(bold=True)
ws3.cell(row=row, column=4, value=grand_expenses)
ws3.cell(row=row, column=4).number_format = money_format
ws3.cell(row=row, column=4).font = Font(bold=True)
ws3.cell(row=row, column=5, value=grand_net)
ws3.cell(row=row, column=5).number_format = money_format
ws3.cell(row=row, column=5).font = Font(bold=True, size=12)

# Notes section
row += 3
ws3.cell(row=row, column=1, value="Notes:")
ws3.cell(row=row, column=1).font = Font(bold=True, size=12)
row += 1
ws3.cell(row=row, column=1, value="1. Income data sourced from Airbnb payout emails (tarikp.us@gmail.com) and Geeves DB.")
row += 1
ws3.cell(row=row, column=1, value="2. Expense data sourced from Amazon, Walmart, and Venmo emails (tariqp@gmail.com).")
row += 1
ws3.cell(row=row, column=1, value="3. General/unallocated expenses split evenly across US properties.")
row += 1
ws3.cell(row=row, column=1, value="4. Tax Pass-Through (TOT) amounts are collected on behalf of local tax authorities and may not be taxable income.")
row += 1
ws3.cell(row=row, column=1, value="5. Some Vrbo bookings from iCal sync have no financial data — payout amounts not available in email scrape.")
row += 1
ws3.cell(row=row, column=1, value="6. Consult your CPA for proper Schedule E reporting and depreciation calculations.")

auto_width(ws3)

# ============================================================
# SHEET 4: Jamaica Income
# ============================================================
ws4 = wb.create_sheet("Jamaica Income")

ws4['A1'] = "2025 Jamaica Property Income"
ws4['A1'].font = Font(bold=True, size=16)
ws4['A2'] = "Property: The Artiste's Boutique (Bohemian Lodge)"
ws4['A3'] = "Location: Jamaica"

row = 5
jamaica_headers = ['Date', 'Guest', 'Check-In', 'Check-Out', 'Amount (USD)', 'Type', 'Payout ID', 'Conf. Code']
for i, h in enumerate(jamaica_headers):
    ws4.cell(row=row, column=i+1, value=h)
style_header_row(ws4, row, len(jamaica_headers))

jamaica_total = 0
jamaica_rental = 0
for r in sorted(jamaica_income, key=lambda x: x.get('date', '')):
    row += 1
    ws4.cell(row=row, column=1, value=r['date'])
    ws4.cell(row=row, column=2, value=r['guest'])
    ws4.cell(row=row, column=3, value=r['check_in'])
    ws4.cell(row=row, column=4, value=r['check_out'])
    ws4.cell(row=row, column=5, value=r['amount'])
    ws4.cell(row=row, column=5).number_format = money_format
    ws4.cell(row=row, column=6, value=r['type'].replace('_', ' ').title())
    ws4.cell(row=row, column=7, value=r['payout_id'])
    ws4.cell(row=row, column=8, value=r['confirmation_code'])
    jamaica_total += r['amount']
    if r['type'] == 'rental_income':
        jamaica_rental += r['amount']

row += 2
ws4.cell(row=row, column=4, value="Total Rental Income:")
ws4.cell(row=row, column=4).font = Font(bold=True)
ws4.cell(row=row, column=5, value=jamaica_rental)
ws4.cell(row=row, column=5).number_format = money_format
ws4.cell(row=row, column=5).font = Font(bold=True)

row += 1
ws4.cell(row=row, column=4, value="Total (incl. Tax Pass-Through):")
ws4.cell(row=row, column=4).font = Font(bold=True)
ws4.cell(row=row, column=5, value=jamaica_total)
ws4.cell(row=row, column=5).number_format = money_format
ws4.cell(row=row, column=5).font = Font(bold=True)

# Monthly breakdown
row += 3
ws4.cell(row=row, column=1, value="Monthly Breakdown")
ws4.cell(row=row, column=1).font = Font(bold=True, size=14)
row += 1
ws4.cell(row=row, column=1, value="Month")
ws4.cell(row=row, column=2, value="Rental Income")
ws4.cell(row=row, column=3, value="Tax Pass-Through")
ws4.cell(row=row, column=4, value="Total")
style_header_row(ws4, row, 4)

for m_idx in range(1, 13):
    m_rental = sum(r['amount'] for r in jamaica_income if r['month'] == m_idx and r['type'] == 'rental_income')
    m_tot = sum(r['amount'] for r in jamaica_income if r['month'] == m_idx and r['type'] == 'tax_passthrough')
    if m_rental > 0 or m_tot > 0:
        row += 1
        ws4.cell(row=row, column=1, value=months[m_idx-1])
        ws4.cell(row=row, column=2, value=m_rental)
        ws4.cell(row=row, column=2).number_format = money_format
        ws4.cell(row=row, column=3, value=m_tot)
        ws4.cell(row=row, column=3).number_format = money_format
        ws4.cell(row=row, column=4, value=m_rental + m_tot)
        ws4.cell(row=row, column=4).number_format = money_format

auto_width(ws4)

# ============================================================
# SHEET 5: All Airbnb Payouts Detail
# ============================================================
ws5 = wb.create_sheet("Airbnb Payouts Detail")

ws5['A1'] = "2025 Airbnb Payout Details (All)"
ws5['A1'].font = Font(bold=True, size=16)
ws5['A2'] = "Source: tarikp.us@gmail.com payout notifications"

row = 4
detail_headers = ['Payout Date', 'Payout ID', 'Total Amount', 'Guest', 'Property', 
                  'Check-In', 'Check-Out', 'Line Amount', 'Type', 'Listing ID']
for i, h in enumerate(detail_headers):
    ws5.cell(row=row, column=i+1, value=h)
style_header_row(ws5, row, len(detail_headers))

for payout in sorted(airbnb_payouts, key=lambda x: x.get('payout_date', '')):
    if 'error' in payout:
        continue
    for item in payout.get('line_items', []):
        row += 1
        ws5.cell(row=row, column=1, value=payout.get('payout_date', ''))
        ws5.cell(row=row, column=2, value=payout.get('payout_id', ''))
        ws5.cell(row=row, column=3, value=payout.get('total_amount', 0))
        ws5.cell(row=row, column=3).number_format = money_format
        ws5.cell(row=row, column=4, value=item.get('guest_name', ''))
        ws5.cell(row=row, column=5, value=normalize_property(item.get('property_name', '')))
        ws5.cell(row=row, column=6, value=item.get('check_in', ''))
        ws5.cell(row=row, column=7, value=item.get('check_out', ''))
        ws5.cell(row=row, column=8, value=item.get('amount', 0))
        ws5.cell(row=row, column=8).number_format = money_format
        ws5.cell(row=row, column=9, value=item.get('description', ''))
        ws5.cell(row=row, column=10, value=item.get('listing_id', ''))

auto_width(ws5)

# ============================================================
# SAVE WORKBOOK
# ============================================================
output_path = '/home/ubuntu/tax_prep/Tarik_Perkins_2025_Property_Tax_Workbook.xlsx'
wb.save(output_path)
print(f"\n✅ Workbook saved to: {output_path}")

# Final summary
print(f"\n{'='*60}")
print(f"2025 PROPERTY TAX WORKBOOK SUMMARY")
print(f"{'='*60}")
print(f"\nUS Properties:")
print(f"  Total Rental Income: ${grand_income:,.2f}")
print(f"  Tax Pass-Through:    ${grand_tot:,.2f}")
print(f"  Total Expenses:      ${grand_expenses:,.2f}")
print(f"  Net Profit/(Loss):   ${grand_net:,.2f}")
print(f"\nJamaica (The Artiste's Boutique):")
print(f"  Total Rental Income: ${jamaica_rental:,.2f}")
print(f"  Tax Pass-Through:    ${jamaica_total - jamaica_rental:,.2f}")
print(f"\nSheets:")
print(f"  1. US Income Dashboard - Monthly pivot + detailed bookings")
print(f"  2. US Expenses - 33 property-related expenses")
print(f"  3. US P&L Summary - Schedule E ready")
print(f"  4. Jamaica Income - Artiste's Boutique bookings")
print(f"  5. Airbnb Payouts Detail - Raw payout line items")
