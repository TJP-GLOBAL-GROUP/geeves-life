#!/usr/bin/env python3
"""
Import Artiste's Boutique expense history from _JamaicanProperties2024.xlsx
into financial_transactions and property_expense_records tables.
"""

import os
import re
import sys
from pathlib import Path
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    os.system("sudo pip3 install openpyxl -q")
    import openpyxl

try:
    import pymysql
except ImportError:
    os.system("sudo pip3 install pymysql -q")
    import pymysql

UPLOAD_DIR = Path("/home/ubuntu/upload")
XLSX_FILE = UPLOAD_DIR / "_JamaicanProperties2024.xlsx"
HOUSEHOLD_ID = "V8lk3KJatvxBTWURf4uo9"

# Account IDs from financial_accounts table
# 1 = Scotia Chequing #620505 (CAD)
# 2 = Scotia Gold Mastercard ****6376 (JMD)
# 3 = Scotia Aero Platinum Mastercard ****6138 (JMD)
# 4 = Airbnb Savings 0108 (USD)

# Paid From mapping to account IDs
PAID_FROM_MAP = {
    "TARIK": None,       # Generic Tarik - no specific account, use NULL
    "TAZ": None,         # Also Tarik, no specific account
    "CASH": None,        # Cash - no account
    "MANNY": None,       # Holdover, skip or use NULL
    "HERMA": None,       # Mother - no account in DB
    "CBNA": None,        # CBNA Checking - not in DB yet
}

def get_db_connection():
    db_url = os.environ.get("DATABASE_URL", "")
    m = re.match(r'mysql2?://([^:]+):([^@]+)@([^:/]+):?(\d+)?/([^?]+)', db_url)
    if not m:
        print("ERROR: Could not parse DATABASE_URL")
        sys.exit(1)
    conn = pymysql.connect(
        host=m.group(3), port=int(m.group(4) or 3306),
        user=m.group(1), password=m.group(2), database=m.group(5),
        ssl={"ssl_ca": None, "check_hostname": False, "verify_mode": 0},
        connect_timeout=30,
    )
    return conn

def parse_date_cell(val):
    """Parse a date value from an Excel cell."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d")
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        # Try various formats
        for fmt in ["%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d %b %Y"]:
            try:
                return datetime.strptime(val, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None

def parse_amount_cell(val):
    """Parse an amount from an Excel cell."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        val = val.strip().replace(",", "").replace("$", "").replace("JMD", "").strip()
        if not val or val in ["-", "N/A", "n/a"]:
            return None
        try:
            return float(val)
        except ValueError:
            return None
    return None

def clean_str(val):
    """Clean a string cell value."""
    if val is None:
        return None
    return str(val).strip() or None

def get_vertical_for_expense():
    """All expenses from this spreadsheet are for Artiste's Boutique."""
    return "artistes_boutique"

def main():
    print(f"Loading spreadsheet: {XLSX_FILE}")
    wb = openpyxl.load_workbook(XLSX_FILE, data_only=True)
    
    print(f"Sheets: {wb.sheetnames}")
    
    # Find the expenses sheet
    expenses_sheet = None
    for name in wb.sheetnames:
        if "expense" in name.lower() or name.lower() in ["expenses", "sheet1", "data"]:
            expenses_sheet = wb[name]
            print(f"Using sheet: {name}")
            break
    
    if not expenses_sheet:
        # Use first sheet
        expenses_sheet = wb.active
        print(f"Using active sheet: {expenses_sheet.title}")
    
    # Read header row
    headers = []
    for cell in expenses_sheet[1]:
        headers.append(clean_str(cell.value) or "")
    
    print(f"Headers: {headers}")
    
    # Map header names to column indices
    col_map = {}
    for i, h in enumerate(headers):
        h_lower = h.lower().strip()
        if "date" in h_lower and "date" not in col_map:
            col_map["date"] = i
        elif h_lower in ["expense", "description", "detail", "item", "name"]:
            if "description" not in col_map:
                col_map["description"] = i
        elif "description" in h_lower or "detail" in h_lower:
            if "description" not in col_map:
                col_map["description"] = i
        elif "amount" in h_lower or "cost" in h_lower or "total" in h_lower:
            if "amount" not in col_map:
                col_map["amount"] = i
        elif "category" in h_lower or "type" in h_lower:
            if "category" not in col_map:
                col_map["category"] = i
        elif h_lower == "paid to":
            col_map["paid_to"] = i
        elif h_lower == "paid from" or ("paid" in h_lower and "from" in h_lower):
            if "paid_from" not in col_map:
                col_map["paid_from"] = i
        elif "note" in h_lower or "comment" in h_lower or "remark" in h_lower:
            if "notes" not in col_map:
                col_map["notes"] = i
        elif "vendor" in h_lower or "supplier" in h_lower or "payee" in h_lower:
            if "vendor" not in col_map:
                col_map["vendor"] = i
        elif "receipt" in h_lower or "invoice" in h_lower or "ref" in h_lower:
            if "reference" not in col_map:
                col_map["reference"] = i
    
    print(f"Column mapping: {col_map}")
    
    # If we couldn't map columns well, print first few rows to debug
    if "date" not in col_map or "amount" not in col_map:
        print("WARNING: Could not find date/amount columns. Printing first 5 rows:")
        for row_idx, row in enumerate(expenses_sheet.iter_rows(min_row=1, max_row=6, values_only=True)):
            print(f"  Row {row_idx+1}: {row}")
        
        # Try to auto-detect by scanning rows
        print("\nAttempting manual column detection...")
        for row_idx, row in enumerate(expenses_sheet.iter_rows(min_row=2, max_row=10, values_only=True)):
            for col_idx, val in enumerate(row):
                if isinstance(val, (datetime, date)):
                    col_map["date"] = col_idx
                    print(f"  Found date at col {col_idx}: {val}")
                    break
            if "date" in col_map:
                break
    
    print(f"\nFinal column mapping: {col_map}")
    
    # Connect to DB
    print("\nConnecting to database...")
    conn = get_db_connection()
    print("Connected!")
    cursor = conn.cursor()
    
    # Get the artistes_boutique property ID
    cursor.execute("SELECT id FROM properties WHERE householdId = %s AND name LIKE %s", 
                   (HOUSEHOLD_ID, "%Artiste%"))
    prop_row = cursor.fetchone()
    property_id = prop_row[0] if prop_row else None
    print(f"Artiste's Boutique property ID: {property_id}")
    
    # Get a default account ID for JMD expenses (use Gold Mastercard as default)
    default_account_id = 2  # Scotia Gold Mastercard ****6376 (JMD)
    
    total_rows = 0
    inserted = 0
    skipped = 0
    
    for row_idx, row in enumerate(expenses_sheet.iter_rows(min_row=2, values_only=True)):
        # Skip empty rows
        if all(v is None for v in row):
            continue
        
        total_rows += 1
        
        # Extract values
        date_val = row[col_map["date"]] if "date" in col_map else None
        amount_val = row[col_map["amount"]] if "amount" in col_map else None
        description_val = row[col_map["description"]] if "description" in col_map else None
        category_val = row[col_map["category"]] if "category" in col_map else None
        paid_from_val = row[col_map["paid_from"]] if "paid_from" in col_map else None
        paid_to_val = row[col_map["paid_to"]] if "paid_to" in col_map else None
        notes_val = row[col_map["notes"]] if "notes" in col_map else None
        vendor_val = row[col_map["vendor"]] if "vendor" in col_map else None
        reference_val = row[col_map["reference"]] if "reference" in col_map else None
        
        date_str = parse_date_cell(date_val)
        amount = parse_amount_cell(amount_val)
        description = clean_str(description_val)
        category = clean_str(category_val)
        paid_from = clean_str(paid_from_val)
        paid_to = clean_str(paid_to_val) if 'paid_to_val' in dir() else None
        notes = clean_str(notes_val)
        vendor = clean_str(vendor_val)
        reference = clean_str(reference_val)
        
        # Skip rows without a date or amount
        if not date_str or amount is None:
            skipped += 1
            continue
        
        # Skip rows with zero amount
        if amount == 0:
            skipped += 1
            continue
        
        # Build description
        desc_parts = []
        if description:
            desc_parts.append(description)
        if vendor:
            desc_parts.append(f"({vendor})")
        if category:
            desc_parts.append(f"[{category}]")
        full_description = " ".join(desc_parts) if desc_parts else f"Artiste's Boutique Expense {date_str}"
        
        # Determine year
        try:
            year = int(date_str[:4])
        except:
            year = None
        
        # Build notes
        notes_parts = []
        if paid_from:
            notes_parts.append(f"Paid from: {paid_from}")
        if notes:
            notes_parts.append(notes)
        full_notes = " | ".join(notes_parts) if notes_parts else None
        
        try:
            cursor.execute("""
                INSERT INTO financial_transactions (
                    householdId, accountId, transactionDate, description,
                    debitAmount, currency, vertical,
                    source, statementYear, referenceNumber,
                    expenseCategory, notes, createdAt
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
            """, (
                HOUSEHOLD_ID,
                default_account_id,
                date_str,
                full_description,
                abs(amount),
                "JMD",
                "artistes_boutique",
                "manual",
                year,
                reference or None,
                category or None,
                full_notes,
            ))
            inserted += 1
            
            # Also insert into property_expense_records
            try:
                month_val = None
                if date_str:
                    try: month_val = int(date_str[5:7])
                    except: pass
                cursor.execute("""
                    INSERT INTO property_expense_records (
                        householdId, property, expenseDate, expenseYear, expenseMonth,
                        expenseDescription, category, amountJMD,
                        paidTo, paidFrom, notes, createdAt
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                """, (
                    HOUSEHOLD_ID,
                    "artistes_boutique",
                    date_str,
                    year,
                    month_val,
                    full_description,
                    category or "UNCATEGORIZED",
                    abs(amount),
                    paid_to if 'paid_to' in dir() and paid_to else None,
                    paid_from or None,
                    full_notes,
                ))
            except Exception as e2:
                print(f"  Warning: property_expense_records insert failed row {row_idx+2}: {e2}")
            
        except Exception as e:
            print(f"  Error inserting row {row_idx+2}: {e} | date={date_str} amount={amount} desc={description}")
            skipped += 1
    
    conn.commit()
    conn.close()
    
    print(f"\n=== EXPENSE IMPORT DONE ===")
    print(f"Total rows processed: {total_rows}")
    print(f"Inserted: {inserted}")
    print(f"Skipped: {skipped}")

if __name__ == "__main__":
    main()
