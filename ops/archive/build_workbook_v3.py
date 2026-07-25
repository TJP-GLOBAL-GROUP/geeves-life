"""
Build Tax Workbook v3 - Uses DB-exported booking data as primary source
Includes: US Income, Jamaica Income, Expenses, P&L, Cancelled bookings
"""
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# Load data
with open('/home/ubuntu/tax_prep/db_2025_bookings.json') as f:
    all_bookings = json.load(f)

with open('/home/ubuntu/tax_prep/parsed/property_expenses_parsed.json') as f:
    expenses = json.load(f)

# Separate by region
us_properties = ['Sunset Studio', 'Morabeza']
jamaica_properties = ["The Artiste's Boutique"]

us_bookings = [b for b in all_bookings if b['propertyName'] in us_properties]
jamaica_bookings = [b for b in all_bookings if b['propertyName'] in jamaica_properties]

# Filter active (non-cancelled) bookings with financials for tax purposes
us_active = [b for b in us_bookings if b.get('bookingStatus') != 'cancelled']
us_cancelled = [b for b in us_bookings if b.get('bookingStatus') == 'cancelled']
jamaica_active = [b for b in jamaica_bookings if b.get('bookingStatus') != 'cancelled']
jamaica_cancelled = [b for b in jamaica_bookings if b.get('bookingStatus') == 'cancelled']

# Styles
header_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=14)
money_format = '#,##0.00'
date_format = 'YYYY-MM-DD'
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font_white = Font(bold=True, size=11, color='FFFFFF')
cancelled_fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
subtotal_fill = PatternFill(start_color='E8F4E8', end_color='E8F4E8', fill_type='solid')

def ts_to_date(ts):
    if not ts or ts == 0:
        return ''
    return datetime.utcfromtimestamp(ts / 1000).strftime('%Y-%m-%d')

def nights(checkin, checkout):
    if not checkin or not checkout:
        return 0
    ci = datetime.utcfromtimestamp(checkin / 1000)
    co = datetime.utcfromtimestamp(checkout / 1000)
    return max((co - ci).days, 0)

def write_header(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    return row + 1

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

# Create workbook
wb = Workbook()

# ============ SHEET 1: US INCOME SUMMARY ============
ws = wb.active
ws.title = "US Income Summary"

ws.cell(row=1, column=1, value="2025 US Rental Income Summary").font = title_font
ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
ws.cell(row=3, column=1, value="Properties: Morabeza + Sunset Studio (4693 SR 414, Burdett, NY 14818)")

# Monthly pivot
row = 5
ws.cell(row=row, column=1, value="Monthly Revenue Summary").font = Font(bold=True, size=12)
row += 1

headers = ['Month', 'Morabeza Revenue', 'Morabeza Net', 'Sunset Studio Revenue', 'Sunset Studio Net', 'Total Revenue', 'Total Net']
row = write_header(ws, row, headers)

months = {}
for b in us_active:
    if b['totalPrice'] > 0 or b['netAmount'] > 0:
        month = ts_to_date(b['checkIn'])[:7]  # YYYY-MM
        if month not in months:
            months[month] = {'Morabeza': {'rev': 0, 'net': 0}, 'Sunset Studio': {'rev': 0, 'net': 0}}
        prop = b['propertyName']
        months[month][prop]['rev'] += b['totalPrice']
        months[month][prop]['net'] += b['netAmount']

total_rev = 0
total_net = 0
for month in sorted(months.keys()):
    m = months[month]
    m_rev = m['Morabeza']['rev']
    m_net = m['Morabeza']['net']
    s_rev = m['Sunset Studio']['rev']
    s_net = m['Sunset Studio']['net']
    t_rev = m_rev + s_rev
    t_net = m_net + s_net
    total_rev += t_rev
    total_net += t_net
    
    ws.cell(row=row, column=1, value=month)
    ws.cell(row=row, column=2, value=m_rev).number_format = money_format
    ws.cell(row=row, column=3, value=m_net).number_format = money_format
    ws.cell(row=row, column=4, value=s_rev).number_format = money_format
    ws.cell(row=row, column=5, value=s_net).number_format = money_format
    ws.cell(row=row, column=6, value=t_rev).number_format = money_format
    ws.cell(row=row, column=7, value=t_net).number_format = money_format
    row += 1

# Totals row
for col in range(1, 8):
    ws.cell(row=row, column=col).fill = subtotal_fill
ws.cell(row=row, column=1, value="TOTAL").font = header_font
ws.cell(row=row, column=6, value=total_rev).number_format = money_format
ws.cell(row=row, column=6).font = header_font
ws.cell(row=row, column=7, value=total_net).number_format = money_format
ws.cell(row=row, column=7).font = header_font

# Summary stats
row += 3
ws.cell(row=row, column=1, value="Summary Statistics").font = Font(bold=True, size=12)
row += 1
us_with_fin = [b for b in us_active if b['totalPrice'] > 0]
total_nights = sum(nights(b['checkIn'], b['checkOut']) for b in us_with_fin)
ws.cell(row=row, column=1, value="Total Confirmed Bookings (with financials):")
ws.cell(row=row, column=2, value=len(us_with_fin))
row += 1
ws.cell(row=row, column=1, value="Total Nights Booked:")
ws.cell(row=row, column=2, value=total_nights)
row += 1
ws.cell(row=row, column=1, value="Average Nightly Rate:")
ws.cell(row=row, column=2, value=total_rev / total_nights if total_nights > 0 else 0).number_format = money_format
row += 1
ws.cell(row=row, column=1, value="Total Gross Revenue:")
ws.cell(row=row, column=2, value=total_rev).number_format = money_format
row += 1
ws.cell(row=row, column=1, value="Total Platform Commission:")
total_comm = sum(b['commissionAmount'] for b in us_with_fin)
ws.cell(row=row, column=2, value=total_comm).number_format = money_format
row += 1
ws.cell(row=row, column=1, value="Total Net Payout:")
ws.cell(row=row, column=2, value=total_net).number_format = money_format
row += 1
ws.cell(row=row, column=1, value="Cancelled Bookings:")
ws.cell(row=row, column=2, value=len(us_cancelled))

auto_width(ws)

# ============ SHEET 2: US BOOKING DETAILS ============
ws2 = wb.create_sheet("US Booking Details")
ws2.cell(row=1, column=1, value="2025 US Booking Details (All Confirmed)").font = title_font

headers = ['Property', 'Guest Name', 'Check-In', 'Check-Out', 'Nights', 'Guests',
           'Confirmation #', 'Total Price', 'Commission', 'Net Payout', 'Cleaning Fee',
           'Currency', 'Platform', 'Data Source']
row = write_header(ws2, 3, headers)

for b in sorted(us_active, key=lambda x: x['checkIn']):
    if b['totalPrice'] > 0 or b['netAmount'] > 0:
        ws2.cell(row=row, column=1, value=b['propertyName'])
        ws2.cell(row=row, column=2, value=b.get('guestName') or 'Unknown')
        ws2.cell(row=row, column=3, value=ts_to_date(b['checkIn']))
        ws2.cell(row=row, column=4, value=ts_to_date(b['checkOut']))
        ws2.cell(row=row, column=5, value=nights(b['checkIn'], b['checkOut']))
        ws2.cell(row=row, column=6, value=b['guestCount'] if b['guestCount'] > 0 else '')
        ws2.cell(row=row, column=7, value=b.get('confirmationNumber') or '')
        ws2.cell(row=row, column=8, value=b['totalPrice']).number_format = money_format
        ws2.cell(row=row, column=9, value=b['commissionAmount']).number_format = money_format
        ws2.cell(row=row, column=10, value=b['netAmount']).number_format = money_format
        ws2.cell(row=row, column=11, value=b['cleaningFee']).number_format = money_format
        ws2.cell(row=row, column=12, value=b.get('currency') or 'USD')
        ws2.cell(row=row, column=13, value=b.get('dataSource') or '')
        ws2.cell(row=row, column=14, value=b.get('dataSource') or '')
        row += 1

# Add iCal-only bookings without financials
row += 1
ws2.cell(row=row, column=1, value="Bookings Without Financial Data (iCal-only):").font = Font(bold=True, italic=True)
row += 1
for b in sorted(us_active, key=lambda x: x['checkIn']):
    if b['totalPrice'] == 0 and b['netAmount'] == 0:
        ws2.cell(row=row, column=1, value=b['propertyName'])
        ws2.cell(row=row, column=2, value=b.get('guestName') or 'Unknown Guest')
        ws2.cell(row=row, column=3, value=ts_to_date(b['checkIn']))
        ws2.cell(row=row, column=4, value=ts_to_date(b['checkOut']))
        ws2.cell(row=row, column=5, value=nights(b['checkIn'], b['checkOut']))
        ws2.cell(row=row, column=7, value=b.get('confirmationNumber') or '')
        ws2.cell(row=row, column=13, value='iCal (no financials)')
        row += 1

auto_width(ws2)

# ============ SHEET 3: US EXPENSES ============
ws3 = wb.create_sheet("US Expenses")
ws3.cell(row=1, column=1, value="2025 US Property Expenses").font = title_font

headers = ['Date', 'Vendor', 'Description', 'Amount', 'Category', 'Likely Property']
row = write_header(ws3, 3, headers)

total_expenses = 0
for e in sorted(expenses, key=lambda x: x.get('date', '')):
    ws3.cell(row=row, column=1, value=e.get('date', ''))
    ws3.cell(row=row, column=2, value=e.get('vendor', ''))
    ws3.cell(row=row, column=3, value=e.get('items_description', ''))
    amt = e.get('total_amount', 0)
    ws3.cell(row=row, column=4, value=amt).number_format = money_format
    ws3.cell(row=row, column=5, value=e.get('category', ''))
    ws3.cell(row=row, column=6, value=e.get('likely_property', 'General'))
    total_expenses += amt
    row += 1

row += 1
ws3.cell(row=row, column=3, value="TOTAL EXPENSES:").font = header_font
ws3.cell(row=row, column=4, value=total_expenses).number_format = money_format
ws3.cell(row=row, column=4).font = header_font

auto_width(ws3)

# ============ SHEET 4: US P&L SUMMARY ============
ws4 = wb.create_sheet("US P&L (Schedule E)")
ws4.cell(row=1, column=1, value="2025 Schedule E - Rental Income & Expenses").font = title_font
ws4.cell(row=2, column=1, value="Tarik Perkins | 4693 SR 414, Burdett, NY 14818")

row = 4
ws4.cell(row=row, column=1, value="INCOME").font = Font(bold=True, size=12)
row += 1

# Income by property
for prop in us_properties:
    prop_bookings = [b for b in us_active if b['propertyName'] == prop and b['totalPrice'] > 0]
    prop_rev = sum(b['totalPrice'] for b in prop_bookings)
    prop_net = sum(b['netAmount'] for b in prop_bookings)
    prop_comm = sum(b['commissionAmount'] for b in prop_bookings)
    
    ws4.cell(row=row, column=1, value=f"  {prop} - Gross Rental Income:")
    ws4.cell(row=row, column=2, value=prop_rev).number_format = money_format
    row += 1
    ws4.cell(row=row, column=1, value=f"  {prop} - Platform Commission:")
    ws4.cell(row=row, column=2, value=-prop_comm).number_format = money_format
    row += 1
    ws4.cell(row=row, column=1, value=f"  {prop} - Net Rental Income:")
    ws4.cell(row=row, column=2, value=prop_net).number_format = money_format
    ws4.cell(row=row, column=2).font = header_font
    row += 1
    row += 1

ws4.cell(row=row, column=1, value="TOTAL NET RENTAL INCOME:").font = header_font
ws4.cell(row=row, column=2, value=total_net).number_format = money_format
ws4.cell(row=row, column=2).font = header_font
row += 2

ws4.cell(row=row, column=1, value="EXPENSES").font = Font(bold=True, size=12)
row += 1

# Expenses by category
by_cat = {}
for e in expenses:
    cat = e.get('category', 'other')
    by_cat[cat] = by_cat.get(cat, 0) + e.get('total_amount', 0)

for cat, amt in sorted(by_cat.items()):
    ws4.cell(row=row, column=1, value=f"  {cat.title()}:")
    ws4.cell(row=row, column=2, value=amt).number_format = money_format
    row += 1

row += 1
ws4.cell(row=row, column=1, value="TOTAL EXPENSES:").font = header_font
ws4.cell(row=row, column=2, value=total_expenses).number_format = money_format
ws4.cell(row=row, column=2).font = header_font
row += 2

net_profit = total_net - total_expenses
ws4.cell(row=row, column=1, value="NET PROFIT (LOSS):").font = Font(bold=True, size=13)
ws4.cell(row=row, column=2, value=net_profit).number_format = money_format
ws4.cell(row=row, column=2).font = Font(bold=True, size=13)

row += 2
ws4.cell(row=row, column=1, value="NOTE: Some bookings have no financial data (iCal-only from Vrbo).").font = Font(italic=True)
row += 1
ws4.cell(row=row, column=1, value="Vrbo payout reports should be pulled directly from the Vrbo host dashboard.").font = Font(italic=True)
row += 1
ws4.cell(row=row, column=1, value=f"Bookings without financials: {len([b for b in us_active if b['totalPrice'] == 0 and b['netAmount'] == 0])}").font = Font(italic=True)

auto_width(ws4)

# ============ SHEET 5: JAMAICA INCOME ============
ws5 = wb.create_sheet("Jamaica Income")
ws5.cell(row=1, column=1, value="2025 Jamaica Rental Income - The Artiste's Boutique").font = title_font
ws5.cell(row=2, column=1, value="7 Dillsbury Ave, Kingston, Jamaica")

# Monthly summary
row = 4
ws5.cell(row=row, column=1, value="Monthly Revenue Summary").font = Font(bold=True, size=12)
row += 1
headers = ['Month', 'Bookings', 'Nights', 'Gross Revenue', 'Commission', 'Net Payout']
row = write_header(ws5, row, headers)

ja_months = {}
for b in jamaica_active:
    if b['totalPrice'] > 0:
        month = ts_to_date(b['checkIn'])[:7]
        if month not in ja_months:
            ja_months[month] = {'count': 0, 'nights': 0, 'rev': 0, 'comm': 0, 'net': 0}
        ja_months[month]['count'] += 1
        ja_months[month]['nights'] += nights(b['checkIn'], b['checkOut'])
        ja_months[month]['rev'] += b['totalPrice']
        ja_months[month]['comm'] += b['commissionAmount']
        ja_months[month]['net'] += b['netAmount']

ja_total_rev = 0
ja_total_net = 0
ja_total_comm = 0
for month in sorted(ja_months.keys()):
    m = ja_months[month]
    ja_total_rev += m['rev']
    ja_total_net += m['net']
    ja_total_comm += m['comm']
    ws5.cell(row=row, column=1, value=month)
    ws5.cell(row=row, column=2, value=m['count'])
    ws5.cell(row=row, column=3, value=m['nights'])
    ws5.cell(row=row, column=4, value=m['rev']).number_format = money_format
    ws5.cell(row=row, column=5, value=m['comm']).number_format = money_format
    ws5.cell(row=row, column=6, value=m['net']).number_format = money_format
    row += 1

for col in range(1, 7):
    ws5.cell(row=row, column=col).fill = subtotal_fill
ws5.cell(row=row, column=1, value="TOTAL").font = header_font
ws5.cell(row=row, column=4, value=ja_total_rev).number_format = money_format
ws5.cell(row=row, column=4).font = header_font
ws5.cell(row=row, column=5, value=ja_total_comm).number_format = money_format
ws5.cell(row=row, column=6, value=ja_total_net).number_format = money_format
ws5.cell(row=row, column=6).font = header_font

auto_width(ws5)

# ============ SHEET 6: JAMAICA BOOKING DETAILS ============
ws6 = wb.create_sheet("Jamaica Booking Details")
ws6.cell(row=1, column=1, value="2025 Jamaica Booking Details - The Artiste's Boutique").font = title_font

headers = ['Guest Name', 'Check-In', 'Check-Out', 'Nights', 'Guests',
           'Confirmation #', 'Total Price', 'Commission', 'Net Payout', 'Cleaning Fee',
           'Currency', 'Status']
row = write_header(ws6, 3, headers)

for b in sorted(jamaica_active + jamaica_cancelled, key=lambda x: x['checkIn']):
    if b['totalPrice'] > 0 or b['netAmount'] > 0:
        is_cancelled = b.get('bookingStatus') == 'cancelled'
        ws6.cell(row=row, column=1, value=b.get('guestName') or 'Unknown')
        ws6.cell(row=row, column=2, value=ts_to_date(b['checkIn']))
        ws6.cell(row=row, column=3, value=ts_to_date(b['checkOut']))
        ws6.cell(row=row, column=4, value=nights(b['checkIn'], b['checkOut']))
        ws6.cell(row=row, column=5, value=b['guestCount'] if b['guestCount'] > 0 else '')
        ws6.cell(row=row, column=6, value=b.get('confirmationNumber') or '')
        ws6.cell(row=row, column=7, value=b['totalPrice']).number_format = money_format
        ws6.cell(row=row, column=8, value=b['commissionAmount']).number_format = money_format
        ws6.cell(row=row, column=9, value=b['netAmount']).number_format = money_format
        ws6.cell(row=row, column=10, value=b['cleaningFee']).number_format = money_format
        ws6.cell(row=row, column=11, value=b.get('currency') or 'USD')
        ws6.cell(row=row, column=12, value='CANCELLED' if is_cancelled else 'Confirmed')
        
        if is_cancelled:
            for col in range(1, 13):
                ws6.cell(row=row, column=col).fill = cancelled_fill
        row += 1

auto_width(ws6)

# ============ SHEET 7: CANCELLED BOOKINGS ============
ws7 = wb.create_sheet("Cancelled Bookings")
ws7.cell(row=1, column=1, value="2025 Cancelled Bookings (All Properties)").font = title_font
ws7.cell(row=2, column=1, value="Note: Cancelled bookings are excluded from income totals")

headers = ['Property', 'Guest Name', 'Check-In', 'Check-Out', 'Confirmation #',
           'Total Price', 'Cancellation Source']
row = write_header(ws7, 4, headers)

all_cancelled = us_cancelled + jamaica_cancelled
for b in sorted(all_cancelled, key=lambda x: x['checkIn']):
    ws7.cell(row=row, column=1, value=b['propertyName'])
    ws7.cell(row=row, column=2, value=b.get('guestName') or 'Unknown')
    ws7.cell(row=row, column=3, value=ts_to_date(b['checkIn']))
    ws7.cell(row=row, column=4, value=ts_to_date(b['checkOut']))
    ws7.cell(row=row, column=5, value=b.get('confirmationNumber') or '')
    ws7.cell(row=row, column=6, value=b['totalPrice']).number_format = money_format
    ws7.cell(row=row, column=7, value=b.get('cancellationSource') or 'Unknown')
    for col in range(1, 8):
        ws7.cell(row=row, column=col).fill = cancelled_fill
    row += 1

row += 1
ws7.cell(row=row, column=1, value=f"Total cancelled: {len(all_cancelled)}")

auto_width(ws7)

# Save
output_path = '/home/ubuntu/tax_prep/Tarik_Perkins_2025_Property_Tax_Workbook_v3.xlsx'
wb.save(output_path)
print(f"Workbook saved to: {output_path}")
print(f"\nSummary:")
print(f"  US Properties: {len(us_with_fin)} bookings with financials, {len(us_cancelled)} cancelled")
print(f"  US Gross Revenue: ${total_rev:.2f}")
print(f"  US Net Payout: ${total_net:.2f}")
print(f"  US Expenses: ${total_expenses:.2f}")
print(f"  US Net Profit: ${net_profit:.2f}")
print(f"  Jamaica: {len([b for b in jamaica_active if b['totalPrice'] > 0])} bookings with financials, {len(jamaica_cancelled)} cancelled")
print(f"  Jamaica Gross Revenue: ${ja_total_rev:.2f}")
print(f"  Jamaica Net Payout: ${ja_total_net:.2f}")
